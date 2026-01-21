from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Body
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_, func, case, or_, cast, String
from datetime import datetime
from utils import get_ist_now, get_ist_date
from decimal import Decimal
from database import get_db
from models import PayrollStructure, Payroll, User, SalaryStructure, PayslipData
from routes.auth import get_current_user
from typing import Optional, Dict, Any
from pydantic import BaseModel
import io

router = APIRouter()

class PayrollStructureCreate(BaseModel):
    name: str
    basic_salary: float
    hra: float = 0
    da: float = 0
    allowances: float = 0
    deductions: float = 0
    tax_percentage: float = 0
    description: Optional[str] = None

class PayrollCreate(BaseModel):
    employee_id: str
    structure_id: int
    month: str
    year: str
    bonus: float = 0
    overtime_hours: float = 0
    deductions: float = 0
    remarks: Optional[str] = None

@router.get("/payroll/structures")
def get_structures(
    db: Session = Depends(get_db)
):
    """Get all payroll structures"""
    # Limit to 200 structures for performance
    structures = db.query(PayrollStructure).limit(200).all()
    return [
        {
            "id": struct.id,
            "name": struct.name,
            "basic_salary": float(struct.basic_salary),
            "hra": float(struct.hra),
            "da": float(struct.da),
            "allowances": float(struct.allowances),
            "deductions": float(struct.deductions),
            "tax_percentage": float(struct.tax_percentage),
            "description": struct.description,
            "created_at": struct.created_at.isoformat() if struct.created_at else None
        }
        for struct in structures
    ]

@router.post("/payroll/structures")
def create_structure(
    structure: PayrollStructureCreate,
    db: Session = Depends(get_db)
):
    """Create a new payroll structure"""
    # Check if structure name already exists
    existing = db.query(PayrollStructure).filter(PayrollStructure.name == structure.name).first()
    if existing:
        raise HTTPException(status_code=400, detail="Structure name already exists")
    
    new_structure = PayrollStructure(
        name=structure.name,
        basic_salary=Decimal(str(structure.basic_salary)),
        hra=Decimal(str(structure.hra)),
        da=Decimal(str(structure.da)),
        allowances=Decimal(str(structure.allowances)),
        deductions=Decimal(str(structure.deductions)),
        tax_percentage=Decimal(str(structure.tax_percentage)),
        description=structure.description
    )
    
    db.add(new_structure)
    db.commit()
    db.refresh(new_structure)
    
    return {
        "message": "Payroll structure created successfully",
        "id": new_structure.id
    }

@router.get("/payroll/list")
def get_payroll_list(
    db: Session = Depends(get_db)
):
    """Get all payroll records"""
    # Limit to 500 payrolls for performance
    payrolls = db.query(Payroll).order_by(Payroll.year.desc(), Payroll.month.desc()).limit(500).all()
    return [
        {
            "id": payroll.id,
            "employee_id": payroll.employee_id,
            "employee_name": payroll.employee_name,
            "structure_id": payroll.structure_id,
            "structure_name": payroll.structure_name,
            "month": payroll.month,
            "year": payroll.year,
            "gross_salary": float(payroll.gross_salary),
            "net_salary": float(payroll.net_salary),
            "status": payroll.status,
            "created_at": payroll.created_at.isoformat() if payroll.created_at else None
        }
        for payroll in payrolls
    ]

@router.post("/payroll/create")
def create_payroll(
    payroll_data: PayrollCreate,
    db: Session = Depends(get_db)
):
    """Create a new payroll record"""
    # Get employee
    employee = db.query(User).filter(User.empid == payroll_data.employee_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Get structure
    structure = db.query(PayrollStructure).filter(PayrollStructure.id == payroll_data.structure_id).first()
    if not structure:
        raise HTTPException(status_code=404, detail="Payroll structure not found")
    
    # Parse month and year
    try:
        month = int(payroll_data.month.split('-')[1])
        year = int(payroll_data.year)
    except:
        raise HTTPException(status_code=400, detail="Invalid month or year format")
    
    # Check if payroll already exists for this employee and period
    existing = db.query(Payroll).filter(
        and_(
            Payroll.employee_id == payroll_data.employee_id,
            Payroll.month == month,
            Payroll.year == year
        )
    ).first()
    
    if existing:
        raise HTTPException(status_code=400, detail="Payroll already exists for this period")
    
    # Calculate salaries
    basic_salary = structure.basic_salary
    hra = structure.hra
    da = structure.da
    allowances = structure.allowances + Decimal(str(payroll_data.bonus))
    overtime_amount = Decimal(str(payroll_data.overtime_hours)) * (basic_salary / (30 * 8))  # Simple calculation
    gross_salary = basic_salary + hra + da + allowances + overtime_amount
    
    # Calculate tax
    tax = gross_salary * (structure.tax_percentage / 100)
    total_deductions = structure.deductions + Decimal(str(payroll_data.deductions)) + tax
    net_salary = gross_salary - total_deductions
    
    new_payroll = Payroll(
        employee_id=payroll_data.employee_id,
        employee_name=employee.name,
        structure_id=structure.id,
        structure_name=structure.name,
        month=month,
        year=year,
        basic_salary=basic_salary,
        hra=hra,
        da=da,
        allowances=allowances,
        bonus=Decimal(str(payroll_data.bonus)),
        overtime_hours=Decimal(str(payroll_data.overtime_hours)),
        overtime_amount=overtime_amount,
        gross_salary=gross_salary,
        tax=tax,
        deductions=total_deductions,
        net_salary=net_salary,
        status='pending',
        remarks=payroll_data.remarks
    )
    
    db.add(new_payroll)
    db.commit()
    db.refresh(new_payroll)
    
    return {
        "message": "Payroll created successfully",
        "id": new_payroll.id
    }

@router.get("/payroll/salary")
def get_salary(
    month: Optional[str] = None,
    year: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Get salary records"""
    query = db.query(Payroll)
    
    if month and year:
        try:
            month_int = int(month.split('-')[1]) if '-' in month else int(month)
            year_int = int(year)
            query = query.filter(
                and_(
                    Payroll.month == month_int,
                    Payroll.year == year_int
                )
            )
        except:
            pass
    
    # Limit to 500 payrolls for performance
    payrolls = query.order_by(Payroll.year.desc(), Payroll.month.desc()).limit(500).all()
    
    return [
        {
            "id": payroll.id,
            "employee_id": payroll.employee_id,
            "employee_name": payroll.employee_name,
            "month": payroll.month,
            "year": payroll.year,
            "basic_salary": float(payroll.basic_salary),
            "hra": float(payroll.hra),
            "da": float(payroll.da),
            "allowances": float(payroll.allowances),
            "bonus": float(payroll.bonus),
            "gross_salary": float(payroll.gross_salary),
            "tax": float(payroll.tax),
            "deductions": float(payroll.deductions),
            "net_salary": float(payroll.net_salary),
            "status": payroll.status
        }
        for payroll in payrolls
    ]

@router.get("/payroll/payslips")
def get_payslips(
    month: Optional[str] = None,
    year: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Get payslips for a specific month and year"""
    query = db.query(Payroll)
    
    if month and year:
        try:
            month_int = int(month.split('-')[1]) if '-' in month else int(month)
            year_int = int(year)
            query = query.filter(
                and_(
                    Payroll.month == month_int,
                    Payroll.year == year_int
                )
            )
        except:
            pass
    
    # Limit to 500 payslips for performance
    payrolls = query.order_by(Payroll.year.desc(), Payroll.month.desc()).limit(500).all()
    
    return [
        {
            "id": payroll.id,
            "employee_id": payroll.employee_id,
            "employee_name": payroll.employee_name,
            "month": payroll.month,
            "year": payroll.year,
            "basic_salary": float(payroll.basic_salary),
            "hra": float(payroll.hra),
            "da": float(payroll.da),
            "allowances": float(payroll.allowances),
            "bonus": float(payroll.bonus),
            "gross_salary": float(payroll.gross_salary),
            "tax": float(payroll.tax),
            "deductions": float(payroll.deductions),
            "net_salary": float(payroll.net_salary),
            "status": payroll.status
        }
        for payroll in payrolls
    ]

@router.get("/payroll/salary-structure")
def get_salary_structure(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get all salary structure records with employee information"""
    try:
        # Build query for salary structures
        query = db.query(
            SalaryStructure.id,
            SalaryStructure.empid,
            SalaryStructure.name,
            SalaryStructure.doj,
            SalaryStructure.salary_per_annum,
            SalaryStructure.salary_per_month,
            SalaryStructure.basic,
            SalaryStructure.hra,
            SalaryStructure.ca,
            SalaryStructure.ma,
            SalaryStructure.sa,
            SalaryStructure.employee_pf,
            SalaryStructure.employee_esi,
            SalaryStructure.professional_tax,
            SalaryStructure.employer_pf,
            SalaryStructure.employer_esi,
            SalaryStructure.variable_pay,
            SalaryStructure.retension_bonus,
            SalaryStructure.net_salary,
            SalaryStructure.monthly_ctc,
            SalaryStructure.pf_check,
            SalaryStructure.esi_check
        )
        
        # Filter by user role - Employee and Manager see only their own
        if current_user.role in ['Employee', 'Manager']:
            query = query.filter(SalaryStructure.empid == current_user.empid)
        
        # Limit to 500 salary structures for performance
        salary_structures = query.limit(500).all()
        
        # Batch load users to avoid N+1 queries
        empids = [salary.empid for salary in salary_structures if salary.empid]
        users_map = {}
        if empids:
            users = db.query(User).filter(User.empid.in_(empids)).all()
            users_map = {user.empid: user for user in users}
        
        result = []
        for salary in salary_structures:
            # Get user information from map
            user = users_map.get(salary.empid) if salary.empid else None
            
            result.append({
                "id": salary.id,
                "empid": salary.empid,
                "employee_name": user.name if user else (salary.name or "N/A"),
                "employee_email": user.email if user else None,
                "employee_image": user.image_base64 if user else None,
                "employee_designation": getattr(user, 'designation', None) if user else None,
                "doj": salary.doj.isoformat() if salary.doj else None,
                "salary_per_annum": float(salary.salary_per_annum) if salary.salary_per_annum else 0,
                "salary_per_month": float(salary.salary_per_month) if salary.salary_per_month else 0,
                "basic": float(salary.basic) if salary.basic else 0,
                "hra": float(salary.hra) if salary.hra else 0,
                "ca": float(salary.ca) if salary.ca else 0,
                "ma": float(salary.ma) if salary.ma else 0,
                "sa": float(salary.sa) if salary.sa else 0,
                "employee_pf": float(salary.employee_pf) if salary.employee_pf else 0,
                "employee_esi": float(salary.employee_esi) if salary.employee_esi else 0,
                "professional_tax": float(salary.professional_tax) if salary.professional_tax else 0,
                "employer_pf": float(salary.employer_pf) if salary.employer_pf else 0,
                "employer_esi": float(salary.employer_esi) if salary.employer_esi else 0,
                "variable_pay": float(salary.variable_pay) if salary.variable_pay else 0,
                "retension_bonus": float(salary.retension_bonus) if salary.retension_bonus else 0,
                "net_salary": float(salary.net_salary) if salary.net_salary else 0,
                "monthly_ctc": float(salary.monthly_ctc) if salary.monthly_ctc else 0,
                "pf_check": bool(salary.pf_check) if salary.pf_check is not None else False,
                "esi_check": bool(salary.esi_check) if salary.esi_check is not None else False
            })
        
        return result
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"Error fetching salary structure: {error_detail}")
        raise HTTPException(status_code=500, detail=f"Error fetching salary structure: {str(e)}")

@router.get("/payroll/salary-structure/export-excel")
def export_salary_structure_excel(
    db: Session = Depends(get_db)
):
    """Export salary structure data to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        # Get all salary structures - limit to 1000 for export performance
        salary_structures = db.query(
            SalaryStructure.id,
            SalaryStructure.empid,
            SalaryStructure.name,
            SalaryStructure.doj,
            SalaryStructure.salary_per_annum,
            SalaryStructure.salary_per_month,
            SalaryStructure.basic,
            SalaryStructure.hra,
            SalaryStructure.ca,
            SalaryStructure.ma,
            SalaryStructure.sa,
            SalaryStructure.employee_pf,
            SalaryStructure.employee_esi,
            SalaryStructure.professional_tax,
            SalaryStructure.employer_pf,
            SalaryStructure.employer_esi,
            SalaryStructure.variable_pay,
            SalaryStructure.retension_bonus,
            SalaryStructure.net_salary,
            SalaryStructure.monthly_ctc,
            SalaryStructure.pf_check,
            SalaryStructure.esi_check
        ).limit(1000).all()
        
        # Batch load users to avoid N+1 queries
        empids = [salary.empid for salary in salary_structures if salary.empid]
        users_map = {}
        if empids:
            users = db.query(User).filter(User.empid.in_(empids)).all()
            users_map = {user.empid: user for user in users}
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Salary Structure"
        
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            "Emp ID", "Employee Name", "Email", "DOJ", "Salary Per Annum", "Salary Per Month",
            "Basic", "HRA", "CA", "MA", "SA", "Employee PF", "Employee ESI",
            "Professional Tax", "Employer PF", "Employer ESI", "Variable Pay",
            "Retention Bonus", "Net Salary", "Monthly CTC", "PF Check", "ESI Check"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Data rows
        for row_num, salary in enumerate(salary_structures, 2):
            # Get user information from map
            user = users_map.get(salary.empid) if salary.empid else None
            
            row_data = [
                salary.empid or "",
                user.name if user else (salary.name or ""),
                user.email if user else "",
                salary.doj.strftime('%Y-%m-%d') if salary.doj else "",
                float(salary.salary_per_annum) if salary.salary_per_annum else 0,
                float(salary.salary_per_month) if salary.salary_per_month else 0,
                float(salary.basic) if salary.basic else 0,
                float(salary.hra) if salary.hra else 0,
                float(salary.ca) if salary.ca else 0,
                float(salary.ma) if salary.ma else 0,
                float(salary.sa) if salary.sa else 0,
                float(salary.employee_pf) if salary.employee_pf else 0,
                float(salary.employee_esi) if salary.employee_esi else 0,
                float(salary.professional_tax) if salary.professional_tax else 0,
                float(salary.employer_pf) if salary.employer_pf else 0,
                float(salary.employer_esi) if salary.employer_esi else 0,
                float(salary.variable_pay) if salary.variable_pay else 0,
                float(salary.retension_bonus) if salary.retension_bonus else 0,
                float(salary.net_salary) if salary.net_salary else 0,
                float(salary.monthly_ctc) if salary.monthly_ctc else 0,
                "Yes" if salary.pf_check else "No",
                "Yes" if salary.esi_check else "No"
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                if isinstance(value, (int, float)) and col_num > 4:  # Numeric columns (after DOJ)
                    cell.number_format = '#,##0'
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"salary_structure_{get_ist_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"Error exporting Excel: {error_detail}")
        raise HTTPException(status_code=500, detail=f"Error exporting Excel: {str(e)}")

@router.post("/payroll/salary-structure/upload-excel")
def upload_salary_structure_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """Upload and update salary structure from Excel file"""
    try:
        from openpyxl import load_workbook
        
        # Read file content
        contents = file.file.read()
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        updated_count = 0
        created_count = 0
        errors = []
        
        # Skip header row, start from row 2
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                if not row[0]:  # Skip empty rows
                    continue
                
                empid = str(row[0]) if row[0] else None
                name = str(row[1]) if row[1] else None
                email = str(row[2]) if row[2] else None
                doj_str = str(row[3]) if row[3] else None
                
                # Parse date
                doj = None
                if doj_str:
                    try:
                        doj = datetime.strptime(doj_str, '%Y-%m-%d').date()
                    except:
                        try:
                            doj = datetime.strptime(doj_str, '%d-%m-%Y').date()
                        except:
                            pass
                
                # Check if empid already exists - update if exists, insert otherwise
                if empid:
                    existing_salary = db.query(SalaryStructure).filter(SalaryStructure.empid == empid).first()
                    
                    # Helper function to convert null/empty to 0
                    def safe_decimal(value):
                        if value is None or (isinstance(value, str) and not value.strip()):
                            return Decimal(0)
                        try:
                            return Decimal(str(value))
                        except:
                            return Decimal(0)
                    
                    if existing_salary:
                        # Update existing record
                        existing_salary.name = name if name else existing_salary.name
                        existing_salary.doj = doj if doj else existing_salary.doj
                        existing_salary.salary_per_annum = safe_decimal(row[4])
                        existing_salary.salary_per_month = safe_decimal(row[5])
                        existing_salary.basic = safe_decimal(row[6])
                        existing_salary.hra = safe_decimal(row[7])
                        existing_salary.ca = safe_decimal(row[8])
                        existing_salary.ma = safe_decimal(row[9])
                        existing_salary.sa = safe_decimal(row[10])
                        existing_salary.employee_pf = safe_decimal(row[11])
                        existing_salary.employee_esi = safe_decimal(row[12])
                        existing_salary.professional_tax = safe_decimal(row[13])
                        existing_salary.employer_pf = safe_decimal(row[14])
                        existing_salary.employer_esi = safe_decimal(row[15])
                        existing_salary.variable_pay = safe_decimal(row[16])
                        existing_salary.retension_bonus = safe_decimal(row[17])
                        existing_salary.net_salary = safe_decimal(row[18])
                        existing_salary.monthly_ctc = safe_decimal(row[19])
                        existing_salary.pf_check = 1 if str(row[20]).lower() in ['yes', '1', 'true'] else 0
                        existing_salary.esi_check = 1 if str(row[21]).lower() in ['yes', '1', 'true'] else 0
                        updated_count += 1
                    else:
                        # Create new record
                        salary = SalaryStructure(empid=empid)
                        salary.name = name
                        salary.doj = doj
                        salary.salary_per_annum = safe_decimal(row[4])
                        salary.salary_per_month = safe_decimal(row[5])
                        salary.basic = safe_decimal(row[6])
                        salary.hra = safe_decimal(row[7])
                        salary.ca = safe_decimal(row[8])
                        salary.ma = safe_decimal(row[9])
                        salary.sa = safe_decimal(row[10])
                        salary.employee_pf = safe_decimal(row[11])
                        salary.employee_esi = safe_decimal(row[12])
                        salary.professional_tax = safe_decimal(row[13])
                        salary.employer_pf = safe_decimal(row[14])
                        salary.employer_esi = safe_decimal(row[15])
                        salary.variable_pay = safe_decimal(row[16])
                        salary.retension_bonus = safe_decimal(row[17])
                        salary.net_salary = safe_decimal(row[18])
                        salary.monthly_ctc = safe_decimal(row[19])
                        salary.pf_check = 1 if str(row[20]).lower() in ['yes', '1', 'true'] else 0
                        salary.esi_check = 1 if str(row[21]).lower() in ['yes', '1', 'true'] else 0
                        
                        db.add(salary)
                        created_count += 1
                else:
                    errors.append(f"Row {row_num}: Emp ID is required for new records")
                    continue
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        db.commit()
        
        return {
            "message": "Excel file processed successfully",
            "updated": updated_count,
            "created": created_count,
            "errors": errors
        }
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"Error uploading Excel: {error_detail}")
        raise HTTPException(status_code=500, detail=f"Error uploading Excel: {str(e)}")

@router.patch("/payroll/salary-structure/{structure_id}/toggle-pf")
def toggle_pf_check(
    structure_id: int,
    db: Session = Depends(get_db)
):
    """Toggle PF check for a salary structure"""
    try:
        from sqlalchemy import update
        
        # Check if record exists
        salary = db.query(SalaryStructure.id, SalaryStructure.pf_check).filter(
            SalaryStructure.id == structure_id
        ).first()
        
        if not salary:
            raise HTTPException(status_code=404, detail="Salary structure not found")
        
        # Toggle the value
        new_value = 1 if salary.pf_check == 0 else 0
        
        # Update using update statement to avoid loading all columns
        db.execute(
            update(SalaryStructure)
            .where(SalaryStructure.id == structure_id)
            .values(pf_check=new_value)
        )
        db.commit()
        
        return {
            "message": "PF check updated successfully",
            "pf_check": bool(new_value)
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        import traceback
        error_detail = traceback.format_exc()
        print(f"Error updating PF check: {error_detail}")
        raise HTTPException(status_code=500, detail=f"Error updating PF check: {str(e)}")

@router.patch("/payroll/salary-structure/{structure_id}/toggle-esi")
def toggle_esi_check(
    structure_id: int,
    db: Session = Depends(get_db)
):
    """Toggle ESI check for a salary structure"""
    try:
        from sqlalchemy import update
        
        # Check if record exists
        salary = db.query(SalaryStructure.id, SalaryStructure.esi_check).filter(
            SalaryStructure.id == structure_id
        ).first()
        
        if not salary:
            raise HTTPException(status_code=404, detail="Salary structure not found")
        
        # Toggle the value
        new_value = 1 if salary.esi_check == 0 else 0
        
        # Update using update statement to avoid loading all columns
        db.execute(
            update(SalaryStructure)
            .where(SalaryStructure.id == structure_id)
            .values(esi_check=new_value)
        )
        db.commit()
        
        return {
            "message": "ESI check updated successfully",
            "esi_check": bool(new_value)
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        import traceback
        error_detail = traceback.format_exc()
        print(f"Error updating ESI check: {error_detail}")
        raise HTTPException(status_code=500, detail=f"Error updating ESI check: {str(e)}")

@router.put("/payroll/salary-structure")
def update_salary_structure(
    salary_data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Update salary_per_annum in users table by empid"""
    from utils import is_admin_or_hr
    
    print(f"=== Salary Update Request ===")
    print(f"Request data: {salary_data}")
    print(f"Current user: {current_user.empid}, Role: {current_user.role}")
    
    if not is_admin_or_hr(current_user):
        raise HTTPException(status_code=403, detail="Only Admin or HR can update salary")
    
    try:
        empid = salary_data.get('empid')
        if not empid:
            raise HTTPException(status_code=400, detail="empid is required")
        
        print(f"Looking for user with empid: {empid}")
        
        # Find user by empid
        user = db.query(User).filter(User.empid == empid).first()
        
        if not user:
            raise HTTPException(status_code=404, detail=f"User with empid {empid} not found")
        
        print(f"Found user: ID={user.id}, Name={user.name}")
        
        # Check if salary_per_annum column exists in users table
        from sqlalchemy import inspect, text
        inspector = inspect(db.bind)
        user_columns = [col['name'] for col in inspector.get_columns('users')]
        print(f"Columns in users table: {user_columns}")
        
        # Update salary_per_annum in users table
        if 'salary_per_annum' in salary_data and salary_data['salary_per_annum'] is not None:
            try:
                new_value = Decimal(str(salary_data['salary_per_annum']))
                print(f"Updating salary_per_annum to {new_value}")
                
                # Use explicit SQL update to update users table
                from sqlalchemy import update as sql_update
                update_stmt = sql_update(User).where(
                    User.empid == empid
                ).values(salary_per_annum=new_value)
                result = db.execute(update_stmt)
                print(f"SQL update executed. Rows affected: {result.rowcount}")
                
                # Also try ORM update if column exists in model
                if hasattr(user, 'salary_per_annum'):
                    user.salary_per_annum = new_value
                    print(f"ORM update: Value set on user object")
                
                db.flush()
                print(f"Flush successful")
                
            except Exception as e:
                print(f"Error converting/updating salary_per_annum: {e}")
                import traceback
                print(traceback.format_exc())
                raise HTTPException(status_code=400, detail=f"Invalid salary_per_annum value: {str(e)}")
        else:
            print(f"salary_per_annum not in request or is None. Request keys: {salary_data.keys()}, salary_per_annum value: {salary_data.get('salary_per_annum')}")
        
        try:
            db.commit()
            print(f"Commit successful")
            
            # Refresh user to get updated value
            db.refresh(user)
            
            # Verify the update by querying the database again
            verify_user = db.query(User).filter(User.empid == empid).first()
            if verify_user:
                salary_value = getattr(verify_user, 'salary_per_annum', None)
                print(f"Verification query - salary_per_annum in DB: {salary_value}")
                
                # If column doesn't exist in model, query directly
                if salary_value is None:
                    result = db.execute(text("SELECT salary_per_annum FROM users WHERE empid = :empid"), {"empid": empid})
                    row = result.fetchone()
                    if row:
                        print(f"Direct SQL query - salary_per_annum: {row[0]}")
                        salary_value = row[0]
        except Exception as e:
            print(f"Error during commit: {e}")
            import traceback
            print(traceback.format_exc())
            db.rollback()
            raise
        
        # Get the final value
        final_value = getattr(user, 'salary_per_annum', None)
        if final_value is None:
            # Try direct query
            try:
                result = db.execute(text("SELECT salary_per_annum FROM users WHERE empid = :empid"), {"empid": empid})
                row = result.fetchone()
                if row:
                    final_value = row[0]
            except:
                pass
        
        return {
            "message": "Salary updated successfully",
            "empid": empid,
            "salary_per_annum": float(final_value) if final_value else None
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        import traceback
        error_detail = traceback.format_exc()
        print(f"Error updating salary: {error_detail}")
        raise HTTPException(status_code=500, detail=f"Error updating salary: {str(e)}")

# Payslip Data Routes
@router.get("/payslip/months")
def get_payslip_months(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get list of months with payslip data and freeze status"""
    try:
        # Build query - ALL roles only see frozen payslips
        query = db.query(
            PayslipData.month,
            PayslipData.year,
            func.max(case((PayslipData.freaze_status == True, 1), else_=0)).label('freaze_status')
        ).filter(
            PayslipData.freaze_status == True  # Only show months with frozen payslips
        )
        
        # For employees, only show months where they have frozen payslips
        if current_user.role == "Employee":
            try:
                emp_id_int = int(current_user.empid) if current_user.empid else None
                if emp_id_int:
                    query = query.filter(PayslipData.emp_id == emp_id_int)
            except:
                pass
        
        # Group by month and year
        results = query.group_by(
            PayslipData.month,
            PayslipData.year
        ).all()
        
        months = []
        for result in results:
            months.append({
                "month": result.month,
                "year": result.year,
                "freaze_status": bool(result.freaze_status) if result.freaze_status is not None and result.freaze_status > 0 else False
            })
        
        return months
    except Exception as e:
        # If table doesn't exist or query fails, return empty list
        print(f"Error in get_payslip_months: {str(e)}")
        import traceback
        traceback.print_exc()
        # Return empty list instead of raising error to prevent 500
        return []

@router.post("/payslip/toggle-freeze")
def toggle_payslip_freeze(
    month: int,
    year: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Toggle freeze status for all payslips in a given month/year"""
    from utils import is_admin_or_hr
    
    if not is_admin_or_hr(current_user):
        raise HTTPException(status_code=403, detail="Only Admin or HR can toggle freeze status")
    
    # Get all payslips for this month/year
    payslips = db.query(PayslipData).filter(
        PayslipData.month == month,
        PayslipData.year == year
    ).all()
    
    if not payslips:
        raise HTTPException(status_code=404, detail="No payslips found for this month/year")
    
    # Get current freeze status (all should have same status)
    current_status = payslips[0].freaze_status if payslips else False
    new_status = not current_status
    
    # Update all payslips
    for payslip in payslips:
        payslip.freaze_status = new_status
        payslip.updated_date = get_ist_date()
        payslip.updated_by = current_user.name or current_user.empid
    
    db.commit()
    
    return {
        "message": f"Freeze status updated to {new_status}",
        "month": month,
        "year": year,
        "freaze_status": new_status,
        "updated_count": len(payslips)
    }

@router.get("/payslip/list")
def get_payslip_list(
    month: Optional[int] = None,
    year: Optional[int] = None,
    page: int = 1,
    limit: int = 20,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get payslip data with pagination"""
    try:
        # Build query
        query = db.query(PayslipData)
        
        # Apply filters
        if month is not None:
            query = query.filter(PayslipData.month == month)
        if year is not None:
            query = query.filter(PayslipData.year == year)
        
        # Apply search filter
        if search and search.strip():
            search_term = search.strip().lower()
            query = query.filter(
                or_(
                    func.lower(PayslipData.full_name).contains(search_term),
                    cast(PayslipData.emp_id, String).contains(search_term)
                )
            )
        
        # Apply freeze status filter - HR role can see all (frozen and unfrozen), others only frozen
        if current_user.role != "HR":
            query = query.filter(PayslipData.freaze_status == True)
        
        # Role-based filtering for employee data
        if current_user.role == "Employee":
            # Employees can only see their own payslips
            try:
                emp_id_int = int(current_user.empid) if current_user.empid else None
                if emp_id_int:
                    query = query.filter(PayslipData.emp_id == emp_id_int)
            except:
                pass
        # Manager, HR, and Admin can see all frozen payslips (filtered above)
        
        # Get total count
        total_count = query.count()
        
        # Apply pagination
        offset = (page - 1) * limit
        payslips = query.order_by(PayslipData.year.desc(), PayslipData.month.desc()).offset(offset).limit(limit).all()
        
        # Format response
        result = []
        for payslip in payslips:
            earnings = payslip.earnings if isinstance(payslip.earnings, dict) else {}
            deductions = payslip.deductions if isinstance(payslip.deductions, dict) else {}
            
            result.append({
                "payslip_id": payslip.payslip_id,
                "full_name": payslip.full_name or "",
                "emp_id": payslip.emp_id,
                "doj": payslip.doj.strftime('%Y-%m-%d') if payslip.doj else None,
                "month": payslip.month,
                "year": payslip.year,
                "salary_per_month": float(payslip.salary_per_month) if payslip.salary_per_month else 0,
                "salary_per_day": float(payslip.salary_per_day) if payslip.salary_per_day else 0,
                "earned_gross": float(payslip.earned_gross) if payslip.earned_gross else 0,
                "net_salary": float(payslip.net_salary) if payslip.net_salary else 0,
                "basic": float(earnings.get("Basic", 0)) if earnings else 0,
                "hra": float(earnings.get("HRA", 0)) if earnings else 0,
                "ca": float(earnings.get("CA", 0)) if earnings else 0,
                "ma": float(earnings.get("MA", 0)) if earnings else 0,
                "sa": float(earnings.get("SA", 0)) if earnings else 0,
                "pf": float(deductions.get("PF", 0)) if deductions else 0,
                "esi": float(deductions.get("ESI", 0)) if deductions else 0,
                "pt": float(deductions.get("PT", 0)) if deductions else 0,
                "lop": float(deductions.get("LOP", 0)) if deductions else 0,
                "tds": float(deductions.get("TDS", 0)) if deductions else 0,
                "lwf": float(deductions.get("LWF", 0)) if deductions else 0,
                "late_logins": float(deductions.get("LateLogins", 0)) if deductions else 0,
                "late_login_deductions": float(deductions.get("LateLogDeduction", 0)) if deductions else 0,
                "total_deductions": float((deductions.get("PF", 0) or 0) + (deductions.get("ESI", 0) or 0) + (deductions.get("PT", 0) or 0) + (deductions.get("LateLogDeduction", 0) or 0) + (deductions.get("LOP", 0) or 0) + (deductions.get("LWF", 0) or 0)) if deductions else 0,
                "present": float(payslip.present) if payslip.present else 0,
                "absent": float(payslip.absent) if payslip.absent else 0,
                "half_days": float(payslip.half_days) if payslip.half_days else 0,
                "holidays": float(payslip.holidays) if payslip.holidays else 0,
                "wo": float(payslip.wo) if payslip.wo else 0,
                "leaves": float(payslip.leaves) if payslip.leaves else 0,
                "payable_days": float(payslip.payable_days) if payslip.payable_days else 0,
                "arrear_salary": float(payslip.arrear_salary) if payslip.arrear_salary else 0,
                "loan_amount": float(payslip.loan_amount) if payslip.loan_amount else 0,
                "other_deduction": float(payslip.other_deduction) if payslip.other_deduction else 0,
                "designation": payslip.designation or "",
                "gross_salary": float(earnings.get("GrossSalary", 0)) if earnings else 0,
                "company_name": payslip.company_name or "",
                "branch_name": payslip.branch_name or "",
                "department_name": payslip.department_name or "",
                "pf_no": payslip.pf_no or "",
                "esi_no": payslip.esi_no or "",
                "freaze_status": payslip.freaze_status if payslip.freaze_status is not None else False
            })
        
        return {
            "data": result,
            "total": total_count,
            "page": page,
            "limit": limit,
            "total_pages": (total_count + limit - 1) // limit if limit > 0 else 1
        }
    except Exception as e:
        print(f"Error in get_payslip_list: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error fetching payslip data: {str(e)}")

@router.get("/payslip/export-excel")
def export_payslip_excel(
    month: Optional[int] = None,
    year: Optional[int] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export payslip data to Excel"""
    if current_user.role not in ["Admin", "HR"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        # Build query (same logic as get_payslip_list)
        query = db.query(PayslipData)
        
        # Apply freeze status filter - HR role can see all (frozen and unfrozen), others only frozen
        if current_user.role != "HR":
            query = query.filter(PayslipData.freaze_status == True)
        
        if month is not None:
            query = query.filter(PayslipData.month == month)
        if year is not None:
            query = query.filter(PayslipData.year == year)
        
        if search and search.strip():
            search_term = search.strip().lower()
            query = query.filter(
                or_(
                    func.lower(PayslipData.full_name).contains(search_term),
                    cast(PayslipData.emp_id, String).contains(search_term)
                )
            )
        
        payslips = query.order_by(PayslipData.year.desc(), PayslipData.month.desc()).all()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Payslip Data"
        
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            'Full Name', 'Emp ID', 'Designation', 'Salary Per Month', 'Salary Per Day',
            'Basic', 'HRA', 'CA', 'MA', 'SA', 'Gross Salary',
            'PF', 'ESI', 'LOP', 'TDS', 'Late Logins', 'Late Login Deductions',
            'Earned Gross', 'Net Salary',
            'Presents', 'Absents', 'Half Days', 'Holidays', 'WO', 'Leaves', 'Payable Days',
            'Arrear Salary', 'Loan Amount', 'Other Deductions', 'Month', 'Year'
        ]
        
        # Column indices for Month and Year (0-based index in headers array)
        month_col_idx = len(headers) - 1  # Second to last column (Month)
        year_col_idx = len(headers)  # Last column (Year)
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for row_idx, payslip in enumerate(payslips, 2):
            earnings = payslip.earnings if isinstance(payslip.earnings, dict) else {}
            deductions = payslip.deductions if isinstance(payslip.deductions, dict) else {}
            
            row_data = [
                payslip.full_name or "",
                payslip.emp_id or 0,
                payslip.designation or "",
                float(payslip.salary_per_month) if payslip.salary_per_month else 0,
                float(payslip.salary_per_day) if payslip.salary_per_day else 0,
                float(earnings.get("Basic", 0)) if earnings else 0,
                float(earnings.get("HRA", 0)) if earnings else 0,
                float(earnings.get("CA", 0)) if earnings else 0,
                float(earnings.get("MA", 0)) if earnings else 0,
                float(earnings.get("SA", 0)) if earnings else 0,
                float(earnings.get("GrossSalary", 0)) if earnings else 0,
                float(deductions.get("PF", 0)) if deductions else 0,
                float(deductions.get("ESI", 0)) if deductions else 0,
                float(deductions.get("LOP", 0)) if deductions else 0,
                float(deductions.get("TDS", 0)) if deductions else 0,
                float(deductions.get("LateLogins", 0)) if deductions else 0,
                float(deductions.get("LateLogDeduction", 0)) if deductions else 0,
                float(payslip.earned_gross) if payslip.earned_gross else 0,
                float(payslip.net_salary) if payslip.net_salary else 0,
                float(payslip.present) if payslip.present else 0,
                float(payslip.absent) if payslip.absent else 0,
                float(payslip.half_days) if payslip.half_days else 0,
                float(payslip.holidays) if payslip.holidays else 0,
                float(payslip.wo) if payslip.wo else 0,
                float(payslip.leaves) if payslip.leaves else 0,
                float(payslip.payable_days) if payslip.payable_days else 0,
                float(payslip.arrear_salary) if payslip.arrear_salary else 0,
                float(payslip.loan_amount) if payslip.loan_amount else 0,
                float(payslip.other_deduction) if payslip.other_deduction else 0,
                int(payslip.month) if payslip.month else 0,
                int(payslip.year) if payslip.year else 0
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = border
                # Apply number formatting to numeric columns, but not to Month and Year
                if col > 3 and col != month_col_idx and col != year_col_idx:
                    cell.number_format = '#,##0.00'
                elif col == month_col_idx or col == year_col_idx:
                    # Format Month and Year as integers without decimals
                    cell.number_format = '0'
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(1, col).column_letter].width = 15
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"payslip_data_{get_ist_now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        print(f"Error exporting payslip Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error exporting Excel: {str(e)}")

@router.post("/payslip/upload-excel")
def upload_payslip_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Upload payslip data from Excel file (insert/update based on emp_id, month, year)"""
    if current_user.role not in ["Admin", "HR"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    try:
        from openpyxl import load_workbook
        
        # Read file content
        contents = file.file.read()
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        updated_count = 0
        inserted_count = 0
        errors = []
        
        # Skip header row, start from row 2
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                if not row[0] or not row[1]:  # Skip rows without name and emp_id
                    continue
                
                # Helper function to convert null/empty to 0
                def safe_decimal(value):
                    if value is None or (isinstance(value, str) and not value.strip()):
                        return Decimal(0)
                    try:
                        return Decimal(str(value))
                    except:
                        return Decimal(0)
                
                def safe_int(value):
                    if value is None or (isinstance(value, str) and not value.strip()):
                        return 0
                    try:
                        return int(value)
                    except:
                        return 0
                
                # Parse row data (matching Excel column order) - null/empty becomes 0
                full_name = str(row[0]) if row[0] else ""
                emp_id = safe_int(row[1]) if row[1] else None
                designation = str(row[2]) if row[2] else ""
                salary_per_month = safe_decimal(row[3])
                salary_per_day = safe_decimal(row[4])
                basic = safe_decimal(row[5])
                hra = safe_decimal(row[6])
                ca = safe_decimal(row[7])
                ma = safe_decimal(row[8])
                sa = safe_decimal(row[9])
                gross_salary = safe_decimal(row[10])
                pf = safe_decimal(row[11])
                esi = safe_decimal(row[12])
                lop = safe_decimal(row[13])
                tds = safe_decimal(row[14])
                late_logins = safe_decimal(row[15])
                late_login_deductions = safe_decimal(row[16])
                earned_gross = safe_decimal(row[17])
                net_salary = safe_decimal(row[18])
                presents = safe_decimal(row[19])
                absents = safe_decimal(row[20])
                half_days = safe_decimal(row[21])
                holidays = safe_decimal(row[22])
                wo = safe_decimal(row[23])
                leaves = safe_decimal(row[24])
                payable_days = safe_decimal(row[25])
                arrear_salary = safe_decimal(row[26])
                loan_amount = safe_decimal(row[27])
                other_deduction = safe_decimal(row[28])
                month = safe_int(row[29]) if row[29] is not None else None
                year = safe_int(row[30]) if row[30] is not None else None
                
                if not emp_id or not month or not year:
                    errors.append(f"Row {row_num}: Missing emp_id, month, or year")
                    continue
                
                # Check if payslip exists (based on emp_id, month, year)
                existing_payslip = db.query(PayslipData).filter(
                    and_(
                        PayslipData.emp_id == emp_id,
                        PayslipData.month == month,
                        PayslipData.year == year
                    )
                ).first()
                
                # If record exists, calculate net_salary using formula: (netsalary + arrear) - (TDS + LoanAmount + otherdeduction)
                # If record doesn't exist, use net_salary from Excel as is
                if existing_payslip:
                    # Use existing net_salary for calculation
                    existing_net_salary = existing_payslip.net_salary if existing_payslip.net_salary else Decimal(0)
                    # Formula: (netsalary + arrear) - (TDS + LoanAmount + otherdeduction)
                    net_salary = (existing_net_salary + arrear_salary) - (tds + loan_amount + other_deduction)
                else:
                    # For new records, use net_salary from Excel directly
                    pass  # net_salary is already set from row[18]
                
                # Prepare earnings and deductions JSONB
                earnings = {
                    "GrossSalary": float(gross_salary),
                    "Basic": float(basic),
                    "HRA": float(hra),
                    "CA": float(ca),
                    "MA": float(ma),
                    "SA": float(sa)
                }
                
                deductions = {
                    "PF": float(pf),
                    "ESI": float(esi),
                    "LOP": float(lop),
                    "TDS": float(tds),
                    "LateLogins": float(late_logins),
                    "LateLogDeduction": float(late_login_deductions)
                }
                
                if existing_payslip:
                    # Update existing
                    existing_payslip.full_name = full_name
                    existing_payslip.designation = designation
                    existing_payslip.salary_per_month = salary_per_month
                    existing_payslip.salary_per_day = salary_per_day
                    existing_payslip.earnings = earnings
                    existing_payslip.deductions = deductions
                    existing_payslip.earned_gross = earned_gross
                    existing_payslip.net_salary = net_salary
                    existing_payslip.present = presents
                    existing_payslip.absent = absents
                    existing_payslip.half_days = half_days
                    existing_payslip.holidays = holidays
                    existing_payslip.wo = wo
                    existing_payslip.leaves = leaves
                    existing_payslip.payable_days = payable_days
                    existing_payslip.arrear_salary = arrear_salary
                    existing_payslip.loan_amount = loan_amount
                    existing_payslip.other_deduction = other_deduction
                    existing_payslip.updated_date = get_ist_date()
                    existing_payslip.updated_by = current_user.name or current_user.empid
                    updated_count += 1
                else:
                    # Insert new
                    new_payslip = PayslipData(
                        full_name=full_name,
                        emp_id=emp_id,
                        designation=designation,
                        salary_per_month=salary_per_month,
                        salary_per_day=salary_per_day,
                        earnings=earnings,
                        deductions=deductions,
                        earned_gross=earned_gross,
                        net_salary=net_salary,
                        present=presents,
                        absent=absents,
                        half_days=half_days,
                        holidays=holidays,
                        wo=wo,
                        leaves=leaves,
                        payable_days=payable_days,
                        arrear_salary=arrear_salary,
                        loan_amount=loan_amount,
                        other_deduction=other_deduction,
                        month=month,
                        year=year,
                        freaze_status=False,
                        created_date=get_ist_date(),
                        created_by=current_user.name or current_user.empid
                    )
                    db.add(new_payslip)
                    inserted_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
                continue
        
        db.commit()
        
        return {
            "message": "Payslip data uploaded successfully",
            "inserted": inserted_count,
            "updated": updated_count,
            "errors": errors[:10]  # Return first 10 errors
        }
        
    except Exception as e:
        db.rollback()
        print(f"Error uploading payslip Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error uploading Excel: {str(e)}")

@router.post("/payroll/generate")
def generate_payroll(
    request_data: dict,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Generate payroll (alias for payslip/generate to match frontend)
    Accepts month in "YYYY-MM" format from frontend
    """
    from routes.payslip_calculate import generate_payslips
    from routes.payslip_calculate import PayslipCalculationRequest
    
    # Parse month from "YYYY-MM" format
    month_str = request_data.get('month', '')
    if not month_str:
        raise HTTPException(status_code=400, detail="Month is required")
    
    if '-' in month_str:
        parts = month_str.split('-')
        year = int(parts[0])
        month = int(parts[1])
    else:
        raise HTTPException(status_code=400, detail="Invalid month format. Expected YYYY-MM")
    
    # Convert request data to PayslipCalculationRequest
    payslip_request = PayslipCalculationRequest(
        company_id=request_data.get('company_id'),
        branch_id=request_data.get('branch_id'),
        department_id=request_data.get('department_id'),
        employee_id=request_data.get('employee_id'),
        month=month,
        year=year
    )
    
    # Call the payslip generate function
    return generate_payslips(payslip_request, db, current_user)

@router.put("/payslip/bank-details/{empid}")
def update_payslip_bank_details(
    empid: str,
    bank_data: Dict[str, Any] = Body(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Update payslip_data table with bank details for all payslips of an employee"""
    try:
        # Convert empid to integer
        try:
            emp_id_int = int(empid)
        except:
            raise HTTPException(status_code=400, detail="Invalid employee ID")
        
        # Only allow users to update their own payslip data, or Admin/HR can update any
        if current_user.role not in ["Admin", "HR"]:
            if current_user.empid != empid:
                raise HTTPException(status_code=403, detail="Access denied - You can only update your own payslip data")
        
        # Update all payslips for this employee
        payslips = db.query(PayslipData).filter(PayslipData.emp_id == emp_id_int).all()
        
        updated_count = 0
        for payslip in payslips:
            if bank_data.get("bank_name") is not None:
                payslip.bank_name = bank_data["bank_name"]
            if bank_data.get("bank_acc_no") is not None:
                payslip.bank_acc_no = bank_data["bank_acc_no"]
            if bank_data.get("ifsc_code") is not None:
                payslip.ifsc_code = bank_data["ifsc_code"]
            if bank_data.get("pan_no") is not None:
                payslip.pan_no = bank_data["pan_no"]
            if bank_data.get("pf_no") is not None:
                payslip.pf_no = bank_data["pf_no"]
            if bank_data.get("esi_no") is not None:
                payslip.esi_no = bank_data["esi_no"]
            updated_count += 1
        
        db.commit()
        
        return {
            "message": f"Successfully updated {updated_count} payslip(s)",
            "updated_count": updated_count
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"Error updating payslip bank details: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error updating payslip bank details: {str(e)}")

@router.post("/payslip/send-email")
def send_payslip_email(
    to_email: str = Body(...),
    subject: str = Body(...),
    message: str = Body(...),
    month: int = Body(...),
    year: int = Body(...),
    emp_id: int = Body(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Send payslip via email with PDF attachment"""
    try:
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        from email.mime.base import MIMEBase
        from email import encoders
        import io
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
        from reportlab.lib.units import mm
        import requests
        
        # Get payslip data
        payslip = db.query(PayslipData).filter(
            and_(
                PayslipData.emp_id == emp_id,
                PayslipData.month == month,
                PayslipData.year == year
            )
        ).first()
        
        if not payslip:
            raise HTTPException(status_code=404, detail="Payslip not found")
        
        # Email configuration
        from_email = "hrms@brihaspathi.com"
        smtp_server = "smtp.gmail.com"
        smtp_port = 587
        smtp_password = "aakbcohigtogpyrl"
        
        # Create email message
        msg = MIMEMultipart('alternative')
        msg['From'] = from_email
        msg['To'] = to_email
        msg['Subject'] = subject
        
        # Create HTML email body
        months = ['', 'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
        month_name = months[month] if month <= 12 else ''
        
        html_body = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    line-height: 1.6;
                    color: #333;
                    max-width: 600px;
                    margin: 0 auto;
                    padding: 20px;
                    background-color: #f4f4f4;
                }}
                .email-container {{
                    background: white;
                    border-radius: 10px;
                    padding: 30px;
                    box-shadow: 0 2px 10px rgba(0,0,0,0.1);
                }}
                .header {{
                    text-align: center;
                    padding-bottom: 20px;
                    border-bottom: 2px solid #007bff;
                    margin-bottom: 30px;
                }}
                .header h1 {{
                    color: #007bff;
                    margin: 0;
                    font-size: 24px;
                }}
                .content {{
                    margin-bottom: 30px;
                }}
                .content p {{
                    margin: 15px 0;
                    font-size: 16px;
                }}
                .payslip-info {{
                    background: #f8f9fa;
                    padding: 20px;
                    border-radius: 8px;
                    margin: 20px 0;
                }}
                .payslip-info p {{
                    margin: 8px 0;
                    font-size: 14px;
                }}
                .footer {{
                    text-align: center;
                    padding-top: 20px;
                    border-top: 1px solid #ddd;
                    margin-top: 30px;
                    color: #666;
                    font-size: 12px;
                }}
                .button {{
                    display: inline-block;
                    padding: 12px 30px;
                    background-color: #007bff;
                    color: white;
                    text-decoration: none;
                    border-radius: 5px;
                    margin: 20px 0;
                }}
            </style>
        </head>
        <body>
            <div class="email-container">
                <div class="header">
                    <h1>📄 Payslip Notification</h1>
                </div>
                <div class="content">
                    <p>{message.replace(chr(10), '<br>')}</p>
                    <div class="payslip-info">
                        <p><strong>Employee Name:</strong> {payslip.full_name or 'N/A'}</p>
                        <p><strong>Employee ID:</strong> {payslip.emp_id or 'N/A'}</p>
                        <p><strong>Month:</strong> {month_name} {year}</p>
                        <p><strong>Net Salary:</strong> ₹{float(payslip.net_salary) if payslip.net_salary else 0:.2f}</p>
                    </div>
                    <p>Please find your payslip attached to this email in PDF format.</p>
                </div>
                <div class="footer">
                    <p><strong>Brihaspathi Technologies Limited</strong></p>
                    <p>This is an automated email. Please do not reply to this email.</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        # Plain text version
        text_body = f"""
{message}

Employee Name: {payslip.full_name or 'N/A'}
Employee ID: {payslip.emp_id or 'N/A'}
Month: {month_name} {year}
Net Salary: ₹{float(payslip.net_salary) if payslip.net_salary else 0:.2f}

Please find your payslip attached to this email in PDF format.

Best regards,
Brihaspathi Technologies Limited
        """
        
        # Attach both versions
        part1 = MIMEText(text_body, 'plain')
        part2 = MIMEText(html_body, 'html')
        msg.attach(part1)
        msg.attach(part2)
        
        # Generate PDF
        pdf_buffer = io.BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=A4, topMargin=20*mm, bottomMargin=20*mm)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#007bff'),
            spaceAfter=20,
            alignment=1
        )
        elements.append(Paragraph(f"PAYSLIP FOR THE MONTH OF {month_name.upper()} {year}", title_style))
        elements.append(Spacer(1, 20))
        
        # Employee Details
        emp_data = [
            ['NAME OF THE EMPLOYEE:', payslip.full_name or '-'],
            ['EMPLOYEE ID:', str(payslip.emp_id) if payslip.emp_id else '-', 'MONTH:', month_name, 'PF NO:', payslip.pf_no or '-'],
            ['DESIGNATION:', payslip.designation or '-', 'PAID DAYS:', f"{float(payslip.payable_days) if payslip.payable_days else 0:.2f}", 'ESI NO:', payslip.esi_no or '-']
        ]
        
        emp_table = Table(emp_data, colWidths=[80*mm, 50*mm, 40*mm, 30*mm, 40*mm, 30*mm])
        emp_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f1f3f5')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        elements.append(emp_table)
        elements.append(Spacer(1, 15))
        
        # Extract earnings and deductions from JSONB
        earnings_dict = payslip.earnings if isinstance(payslip.earnings, dict) else {}
        deductions_dict = payslip.deductions if isinstance(payslip.deductions, dict) else {}
        
        # Earnings
        elements.append(Paragraph("EARNINGS", ParagraphStyle('SectionHeader', parent=styles['Heading2'], fontSize=14, textColor=colors.white, backColor=colors.HexColor('#6b7785'), alignment=1)))
        earnings_data = [
            ['BASIC', 'HRA', 'CONV', 'ARREARS', 'FIX HRA', 'OTHER ALLOW', 'UNIFORM ALLOW', 'MED ALLOW', 'CCA', 'MOBILE ALLOWANCES'],
            [
                f"{float(earnings_dict.get('Basic', 0)):.2f}",
                f"{float(earnings_dict.get('HRA', 0)):.2f}",
                f"{float(earnings_dict.get('CA', 0)):.2f}",
                f"{float(payslip.arrear_salary) if payslip.arrear_salary else 0:.2f}",
                "0.00",
                f"{float(earnings_dict.get('SA', 0)):.2f}",
                "0.00",
                f"{float(earnings_dict.get('MA', 0)):.2f}",
                "0.00",
                "0.00"
            ]
        ]
        earnings_table = Table(earnings_data, colWidths=[20*mm] * 10)
        earnings_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f1f3f5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        elements.append(earnings_table)
        elements.append(Spacer(1, 15))
        
        # Deductions
        elements.append(Paragraph("DEDUCTIONS", ParagraphStyle('SectionHeader', parent=styles['Heading2'], fontSize=14, textColor=colors.white, backColor=colors.HexColor('#6b7785'), alignment=1)))
        deductions_data = [
            ['PF', 'ESI', 'PROF TAX', 'LWF', 'IT', 'LIC', 'OTHER', 'BANK LOAN', 'COMP LOAN', 'RENT PAID', 'SALARY ADV'],
            [
                f"{float(deductions_dict.get('PF', 0)):.2f}",
                f"{float(deductions_dict.get('ESI', 0)):.2f}",
                f"{float(deductions_dict.get('PT', 0)):.2f}",
                "0.00",
                f"{float(deductions_dict.get('TDS', 0)):.2f}",
                "0.00",
                f"{float(payslip.other_deduction) if payslip.other_deduction else 0:.2f}",
                "0.00",
                f"{float(payslip.loan_amount) if payslip.loan_amount else 0:.2f}",
                "0.00",
                "0.00"
            ]
        ]
        deductions_table = Table(deductions_data, colWidths=[18*mm] * 11)
        deductions_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f1f3f5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        elements.append(deductions_table)
        elements.append(Spacer(1, 15))
        
        # Calculate total deductions
        total_deductions = (
            float(deductions_dict.get('PF', 0)) +
            float(deductions_dict.get('ESI', 0)) +
            float(deductions_dict.get('PT', 0)) +
            float(deductions_dict.get('TDS', 0)) +
            float(payslip.other_deduction if payslip.other_deduction else 0) +
            float(payslip.loan_amount if payslip.loan_amount else 0)
        )
        
        # Totals
        total_data = [
            ['TOTAL EARNINGS (IN INR)', f"{float(payslip.earned_gross) if payslip.earned_gross else 0:.2f}"],
            ['TOTAL DEDUCTIONS (IN INR)', f"{total_deductions:.2f}"],
            ['NET PAY (IN INR)', f"{float(payslip.net_salary) if payslip.net_salary else 0:.2f}"]
        ]
        total_table = Table(total_data, colWidths=[100*mm, 80*mm])
        total_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f1f3f5')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey)
        ]))
        elements.append(total_table)
        elements.append(Spacer(1, 10))
        
        # Footer
        elements.append(Paragraph("** system generated print out. no signature required **", ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=1)))
        
        doc.build(elements)
        pdf_buffer.seek(0)
        
        # Attach PDF
        attachment = MIMEBase('application', 'octet-stream')
        attachment.set_payload(pdf_buffer.read())
        encoders.encode_base64(attachment)
        attachment.add_header('Content-Disposition', f'attachment; filename=Payslip_{month_name}_{year}_{payslip.emp_id}.pdf')
        msg.attach(attachment)
        
        # Send email
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(from_email, smtp_password)
        server.send_message(msg)
        server.quit()
        
        return {
            "success": True,
            "message": "Payslip sent via email successfully"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error sending payslip email: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error sending email: {str(e)}")

