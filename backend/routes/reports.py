from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_
from database import get_db
from models import Project, Task, User, Issue, TaskRating
from schemas import ReportFilter
from routes.auth import get_current_user
from datetime import datetime, date
from typing import Optional
import io

router = APIRouter(prefix="/reports", tags=["Reports"])

@router.get("/filters")
def get_filter_options(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get available filter options for reports"""
    if current_user.role == "Employee":
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Get projects
    project_query = db.query(Project)
    if current_user.role == "Manager":
        project_query = project_query.filter(Project.project_head_id == current_user.id)
    
    projects = [{"id": p.id, "name": p.name} for p in project_query.all()]
    
    # Get employees
    employee_query = db.query(User).filter(User.is_active == True)
    if current_user.role == "Manager":
        employee_query = employee_query.filter(User.report_to_id == current_user.empid)
    else:
        employee_query = employee_query.filter(User.role.in_(["Employee", "Manager"]))
    
    employees = [{"id": e.id, "name": e.name, "empid": e.empid} for e in employee_query.all()]
    
    return {
        "projects": projects,
        "employees": employees
    }

@router.post("/generate")
def generate_report(
    filters: ReportFilter,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Generate report data based on filters"""
    if current_user.role == "Employee":
        raise HTTPException(status_code=403, detail="Access denied")
    
    # Build task query
    task_query = db.query(Task)
    
    if filters.project_id:
        task_query = task_query.filter(Task.project_id == filters.project_id)
    
    if filters.employee_id:
        task_query = task_query.filter(
            or_(
                Task.assigned_to_id == filters.employee_id,
                Task.assigned_by_id == filters.employee_id
            )
        )
    
    if filters.start_date:
        task_query = task_query.filter(Task.created_at >= datetime.combine(filters.start_date, datetime.min.time()))
    
    if filters.end_date:
        task_query = task_query.filter(Task.created_at <= datetime.combine(filters.end_date, datetime.max.time()))
    
    tasks = task_query.all()
    
    # Get project details if filtered
    project = None
    if filters.project_id:
        project = db.query(Project).filter(Project.id == filters.project_id).first()
    
    # Get employee details if filtered
    employee = None
    if filters.employee_id:
        employee = db.query(User).filter(User.id == filters.employee_id).first()
    
    # Calculate stats
    total_tasks = len(tasks)
    completed_tasks = sum(1 for t in tasks if t.status == "done")
    in_progress_tasks = sum(1 for t in tasks if t.status == "in-progress")
    pending_tasks = sum(1 for t in tasks if t.status == "todo")
    
    today = datetime.now().date()
    delayed_tasks = sum(1 for t in tasks if t.due_date and t.due_date < today and t.status != "done")
    
    # Get ratings if employee is filtered
    ratings = []
    avg_rating = 0
    if filters.employee_id:
        ratings_query = db.query(TaskRating).filter(TaskRating.ratee_id == filters.employee_id)
        ratings = ratings_query.all()
        if ratings:
            avg_rating = sum(r.score for r in ratings) / len(ratings)
    
    # Build task details
    task_details = []
    for task in tasks:
        task_details.append({
            "id": task.id,
            "title": task.title,
            "status": task.status,
            "priority": task.priority,
            "assigned_to": task.assigned_to_name,
            "assigned_by": task.assigned_by_name,
            "start_date": task.start_date.isoformat() if task.start_date else None,
            "due_date": task.due_date.isoformat() if task.due_date else None,
            "percent_complete": task.percent_complete,
            "created_at": task.created_at.isoformat()
        })
    
    return {
        "summary": {
            "total_tasks": total_tasks,
            "completed": completed_tasks,
            "in_progress": in_progress_tasks,
            "pending": pending_tasks,
            "delayed": delayed_tasks,
            "completion_rate": round((completed_tasks / total_tasks * 100) if total_tasks > 0 else 0, 2)
        },
        "project": {
            "id": project.id,
            "name": project.name,
            "status": project.status,
            "progress": project.progress_percent
        } if project else None,
        "employee": {
            "id": employee.id,
            "name": employee.name,
            "empid": employee.empid,
            "avg_rating": round(avg_rating, 2),
            "total_ratings": len(ratings)
        } if employee else None,
        "tasks": task_details,
        "generated_at": datetime.now().isoformat()
    }

@router.post("/download/excel")
def download_excel_report(
    filters: ReportFilter,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Download report as Excel file"""
    if current_user.role == "Employee":
        raise HTTPException(status_code=403, detail="Access denied")
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    # Get report data
    report_data = generate_report(filters, db, current_user)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Task Report"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "Task Management Report"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Generated date
    ws['A2'] = f"Generated: {report_data['generated_at']}"
    
    # Summary section
    ws['A4'] = "Summary"
    ws['A4'].font = Font(bold=True, size=12)
    
    summary = report_data['summary']
    ws['A5'] = f"Total Tasks: {summary['total_tasks']}"
    ws['A6'] = f"Completed: {summary['completed']}"
    ws['A7'] = f"In Progress: {summary['in_progress']}"
    ws['A8'] = f"Pending: {summary['pending']}"
    ws['A9'] = f"Delayed: {summary['delayed']}"
    ws['A10'] = f"Completion Rate: {summary['completion_rate']}%"
    
    # Task details header
    headers = ['ID', 'Title', 'Status', 'Priority', 'Assigned To', 'Assigned By', 'Start Date', 'Due Date', '% Complete']
    row_num = 12
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Task data
    for task in report_data['tasks']:
        row_num += 1
        ws.cell(row=row_num, column=1, value=task['id']).border = thin_border
        ws.cell(row=row_num, column=2, value=task['title']).border = thin_border
        ws.cell(row=row_num, column=3, value=task['status']).border = thin_border
        ws.cell(row=row_num, column=4, value=task['priority']).border = thin_border
        ws.cell(row=row_num, column=5, value=task['assigned_to'] or '').border = thin_border
        ws.cell(row=row_num, column=6, value=task['assigned_by'] or '').border = thin_border
        ws.cell(row=row_num, column=7, value=task['start_date'] or '').border = thin_border
        ws.cell(row=row_num, column=8, value=task['due_date'] or '').border = thin_border
        ws.cell(row=row_num, column=9, value=task['percent_complete']).border = thin_border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"task_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.post("/download/pdf")
def download_pdf_report(
    filters: ReportFilter,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Download report as PDF file"""
    if current_user.role == "Employee":
        raise HTTPException(status_code=403, detail="Access denied")
    
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    
    # Get report data
    report_data = generate_report(filters, db, current_user)
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=20,
        alignment=1
    )
    elements.append(Paragraph("Task Management Report", title_style))
    elements.append(Paragraph(f"Generated: {report_data['generated_at']}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Summary
    elements.append(Paragraph("Summary", styles['Heading2']))
    summary = report_data['summary']
    summary_data = [
        ["Total Tasks", str(summary['total_tasks'])],
        ["Completed", str(summary['completed'])],
        ["In Progress", str(summary['in_progress'])],
        ["Pending", str(summary['pending'])],
        ["Delayed", str(summary['delayed'])],
        ["Completion Rate", f"{summary['completion_rate']}%"]
    ]
    
    summary_table = Table(summary_data, colWidths=[150, 100])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('PADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Tasks table
    elements.append(Paragraph("Task Details", styles['Heading2']))
    
    task_headers = ['ID', 'Title', 'Status', 'Priority', 'Assigned To', '% Complete']
    task_data = [task_headers]
    
    for task in report_data['tasks'][:50]:  # Limit to 50 tasks for PDF
        task_data.append([
            str(task['id']),
            task['title'][:30] + '...' if len(task['title']) > 30 else task['title'],
            task['status'],
            task['priority'],
            task['assigned_to'] or '-',
            str(task['percent_complete'])
        ])
    
    task_table = Table(task_data, colWidths=[30, 150, 60, 60, 80, 50])
    task_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('PADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(task_table)
    
    doc.build(elements)
    output.seek(0)
    
    filename = f"task_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )






