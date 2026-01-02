from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.orm import Session
from sqlalchemy import extract
from database import get_db
from models import User
from routes.auth import get_current_user
from datetime import date, datetime
from typing import Optional
import io
import zipfile
import base64
from fastapi.responses import StreamingResponse
import json

router = APIRouter(prefix="/employee-data", tags=["Employee Data"])

@router.get("/export/excel/employee-details")
def export_employee_details_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export employee details to Excel"""
    if current_user.role not in ["Admin", "HR"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    try:
        users = db.query(User).filter(User.is_active == True).order_by(User.empid).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Employee Details"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['Emp ID', 'Name', 'Email', 'Phone', 'Role', 'DOJ', 'DOB']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row_idx, user in enumerate(users, 2):
            ws.cell(row=row_idx, column=1, value=user.empid or '').border = thin_border
            ws.cell(row=row_idx, column=2, value=user.name or '').border = thin_border
            ws.cell(row=row_idx, column=3, value=user.email or '').border = thin_border
            ws.cell(row=row_idx, column=4, value=user.phone or '').border = thin_border
            ws.cell(row=row_idx, column=5, value=user.role or '').border = thin_border
            ws.cell(row=row_idx, column=6, value=user.doj.strftime('%Y-%m-%d') if user.doj else '').border = thin_border
            ws.cell(row=row_idx, column=7, value=user.dob.strftime('%Y-%m-%d') if user.dob else '').border = thin_border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"employee_details_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error exporting employee details: {str(e)}")

@router.get("/export/excel/bank-details")
def export_bank_details_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export bank details to Excel"""
    if current_user.role not in ["Admin", "HR"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    try:
        users = db.query(User).filter(
            User.is_active == True,
            User.bank_details.isnot(None)
        ).order_by(User.empid).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Bank Details"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['Emp ID', 'Name', 'Bank Details']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row_idx, user in enumerate(users, 2):
            bank_details_str = json.dumps(user.bank_details, indent=2) if user.bank_details else ''
            ws.cell(row=row_idx, column=1, value=user.empid or '').border = thin_border
            ws.cell(row=row_idx, column=2, value=user.name or '').border = thin_border
            ws.cell(row=row_idx, column=3, value=bank_details_str).border = thin_border
            ws.cell(row=row_idx, column=3).alignment = Alignment(wrap_text=True, vertical='top')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 60
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"bank_details_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error exporting bank details: {str(e)}")

@router.get("/export/excel/family-details")
def export_family_details_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export family details to Excel"""
    if current_user.role not in ["Admin", "HR"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    try:
        users = db.query(User).filter(
            User.is_active == True,
            User.family_details.isnot(None)
        ).order_by(User.empid).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Family Details"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['Emp ID', 'Name', 'Family Details']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row_idx, user in enumerate(users, 2):
            family_details_str = json.dumps(user.family_details, indent=2) if user.family_details else ''
            ws.cell(row=row_idx, column=1, value=user.empid or '').border = thin_border
            ws.cell(row=row_idx, column=2, value=user.name or '').border = thin_border
            ws.cell(row=row_idx, column=3, value=family_details_str).border = thin_border
            ws.cell(row=row_idx, column=3).alignment = Alignment(wrap_text=True, vertical='top')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 60
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"family_details_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error exporting family details: {str(e)}")

@router.get("/export/excel/nominee-details")
def export_nominee_details_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export nominee details to Excel"""
    if current_user.role not in ["Admin", "HR"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    try:
        users = db.query(User).filter(
            User.is_active == True,
            User.nominee_details.isnot(None)
        ).order_by(User.empid).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Nominee Details"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['Emp ID', 'Name', 'Nominee Details']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row_idx, user in enumerate(users, 2):
            nominee_details_str = json.dumps(user.nominee_details, indent=2) if user.nominee_details else ''
            ws.cell(row=row_idx, column=1, value=user.empid or '').border = thin_border
            ws.cell(row=row_idx, column=2, value=user.name or '').border = thin_border
            ws.cell(row=row_idx, column=3, value=nominee_details_str).border = thin_border
            ws.cell(row=row_idx, column=3).alignment = Alignment(wrap_text=True, vertical='top')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 60
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"nominee_details_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error exporting nominee details: {str(e)}")

@router.get("/export/excel/birthdays")
def export_birthdays_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export all birthdays ordered by month to Excel"""
    if current_user.role not in ["Admin", "HR"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    try:
        users = db.query(User).filter(
            User.is_active == True,
            User.dob.isnot(None)
        ).order_by(extract('month', User.dob), extract('day', User.dob)).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Birthdays"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['Emp ID', 'Name', 'Month', 'Date']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        month_names = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        
        # Data
        for row_idx, user in enumerate(users, 2):
            month = user.dob.month if user.dob else None
            day = user.dob.day if user.dob else None
            ws.cell(row=row_idx, column=1, value=user.empid or '').border = thin_border
            ws.cell(row=row_idx, column=2, value=user.name or '').border = thin_border
            ws.cell(row=row_idx, column=3, value=month_names[month] if month else '').border = thin_border
            ws.cell(row=row_idx, column=4, value=day or '').border = thin_border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"birthdays_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error exporting birthdays: {str(e)}")

@router.get("/export/excel/anniversaries")
def export_anniversaries_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export all anniversaries (DOJ) ordered by month to Excel"""
    if current_user.role not in ["Admin", "HR"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    try:
        users = db.query(User).filter(
            User.is_active == True,
            User.doj.isnot(None)
        ).order_by(extract('month', User.doj), extract('day', User.doj)).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Anniversaries"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['Emp ID', 'Name', 'Month', 'Date']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        month_names = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        
        # Data
        for row_idx, user in enumerate(users, 2):
            month = user.doj.month if user.doj else None
            day = user.doj.day if user.doj else None
            ws.cell(row=row_idx, column=1, value=user.empid or '').border = thin_border
            ws.cell(row=row_idx, column=2, value=user.name or '').border = thin_border
            ws.cell(row=row_idx, column=3, value=month_names[month] if month else '').border = thin_border
            ws.cell(row=row_idx, column=4, value=day or '').border = thin_border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"anniversaries_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error exporting anniversaries: {str(e)}")

@router.post("/upload/documents")
async def upload_documents_folder(
    folder_name: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Upload a zip folder containing images named by empid and update documents JSONB column"""
    if current_user.role not in ["Admin", "HR"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    try:
        # Check if file is a zip
        if not file.filename.endswith('.zip'):
            raise HTTPException(status_code=400, detail="Please upload a ZIP file")
        
        # Read zip file
        zip_content = await file.read()
        zip_file = zipfile.ZipFile(io.BytesIO(zip_content))
        
        updated_count = 0
        skipped_count = 0
        error_count = 0
        
        # Process each file in the zip
        for file_info in zip_file.filelist:
            if file_info.is_dir():
                continue
            
            # Get filename without path and extension
            filename = file_info.filename.split('/')[-1]  # Handle nested paths
            empid = filename.rsplit('.', 1)[0]  # Remove extension
            
            # Find user by empid
            user = db.query(User).filter(User.empid == empid).first()
            
            if not user:
                skipped_count += 1
                continue
            
            try:
                # Read image file
                image_data = zip_file.read(file_info.filename)
                image_base64 = base64.b64encode(image_data).decode('utf-8')
                
                # Determine image format
                file_ext = filename.rsplit('.', 1)[-1].lower()
                if file_ext in ['jpg', 'jpeg']:
                    image_data_uri = f"data:image/jpeg;base64,{image_base64}"
                elif file_ext == 'png':
                    image_data_uri = f"data:image/png;base64,{image_base64}"
                else:
                    image_data_uri = f"data:image/{file_ext};base64,{image_base64}"
                
                # Get existing documents or initialize
                documents = user.documents if user.documents else []
                if not isinstance(documents, list):
                    documents = []
                
                # Check if document with same name exists
                doc_exists = False
                for idx, doc in enumerate(documents):
                    if isinstance(doc, dict) and doc.get('name') == folder_name:
                        # Update existing document
                        documents[idx] = {
                            'name': folder_name,
                            'image': image_data_uri
                        }
                        doc_exists = True
                        break
                
                if not doc_exists:
                    # Add new document
                    documents.append({
                        'name': folder_name,
                        'image': image_data_uri
                    })
                
                # Update user
                user.documents = documents
                db.commit()
                updated_count += 1
                
            except Exception as e:
                error_count += 1
                print(f"Error processing file {filename}: {str(e)}")
                continue
        
        return {
            "message": "Documents uploaded successfully",
            "updated": updated_count,
            "skipped": skipped_count,
            "errors": error_count
        }
        
    except zipfile.BadZipFile:
        raise HTTPException(status_code=400, detail="Invalid ZIP file")
    except Exception as e:
        import traceback
        traceback.print_exc()
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error uploading documents: {str(e)}")

