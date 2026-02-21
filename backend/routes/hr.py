from fastapi import APIRouter, Depends, HTTPException, Body, File, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_, extract, String, cast
from datetime import datetime, date, timedelta
from utils import get_ist_now
from database import get_db
from models import Attendance, User, PunchLog, AttendanceList, WeekOffDate, Holiday, Leave, AttendanceCycle, LeaveBalanceList
from routes.auth import get_current_user
from typing import Optional, List
from pydantic import BaseModel
from config import settings
import io

router = APIRouter()

# Attendance Generation Schema
class AttendanceGenerateRequest(BaseModel):
    month: int
    year: int
    employee_id: Optional[str] = None

@router.get("/attendance/count")
def get_attendance_count(
    date: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Get attendance count statistics for a specific date"""
    if not date:
        date_str = datetime.now().date().isoformat()
    else:
        date_str = date
    
    try:
        target_date = datetime.fromisoformat(date_str).date()
    except:
        target_date = datetime.now().date()
    
    # Get total employees
    total_employees = db.query(User).filter(User.role.in_(['Employee', 'Manager'])).count()
    
    # Get attendance records for the date
    attendance_records = db.query(Attendance).filter(
        Attendance.date == target_date
    ).all()
    
    present_count = sum(1 for a in attendance_records if a.status == 'present')
    absent_count = total_employees - present_count
    leave_count = sum(1 for a in attendance_records if a.status == 'leave')
    late_count = sum(1 for a in attendance_records if a.status == 'late')
    
    return {
        "totalEmployees": total_employees,
        "presentToday": present_count,
        "absentToday": absent_count,
        "onLeave": leave_count,
        "lateArrivals": late_count
    }

@router.get("/attendance/count-details")
def get_attendance_count_details(
    month: int,
    year: int,
    employee_id: Optional[str] = None,
    search: Optional[str] = None,
    page: int = 1,
    limit: int = 20,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get detailed attendance count for all employees for a month"""
    if current_user.role not in ["Admin", "Manager", "HR"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    try:
        query = db.query(AttendanceList).filter(
            and_(
                AttendanceList.month == str(month),
                AttendanceList.year == year
            )
        )
        
        # Apply search filter
        if search and search.strip():
            search_term = search.strip().lower()
            query = query.filter(
                or_(
                    func.lower(AttendanceList.name).contains(search_term),
                    cast(AttendanceList.empid, String).contains(search_term)
                )
            )
        
        if employee_id:
            try:
                empid_int = int(employee_id) if employee_id and str(employee_id).isdigit() else None
                if empid_int:
                    query = query.filter(AttendanceList.empid == empid_int)
            except:
                pass
        
        # Get total count for pagination
        total_count = query.count()
        
        # Apply pagination
        offset = (page - 1) * limit
        records = query.order_by(AttendanceList.name).offset(offset).limit(limit).all()
        
        return {
            "data": [
                {
                    "id": rec.id,
                    "employee_id": rec.empid or getattr(rec, 'employee_id', None),
                    "employee_name": rec.name,
                    "total_days": float(rec.total_days) if rec.total_days else 0,
                    "working_days": float(rec.working_days) if rec.working_days else 0,
                    "week_offs": rec.week_offs or 0,
                    "holidays": float(rec.holi_days) if rec.holi_days else 0,
                    "presents": float(rec.presents) if rec.presents else 0,
                    "absents": float(rec.absents) if rec.absents else 0,
                    "half_days": float(rec.half_days) if rec.half_days else 0,
                    "late_logs": rec.late_logs or 0,
                    "lops": float(rec.lops) if rec.lops else 0,
                    "cl": float(rec.cl) if rec.cl else 0,
                    "sl": float(rec.sl) if rec.sl else 0,
                    "comp_offs": float(rec.comp_offs) if rec.comp_offs else 0,
                    "payble_days": float(rec.payble_days) if rec.payble_days else 0
                }
                for rec in records
            ],
            "pagination": {
                "page": page,
                "limit": limit,
                "total": total_count,
                "total_pages": (total_count + limit - 1) // limit if limit > 0 else 1
            }
        }
    except Exception as e:
        print(f"Error fetching attendance count details: {e}")
        import traceback
        traceback.print_exc()
        # Return empty structure on error to avoid 500 on UI
        return {"data": [], "pagination": {"page": 1, "limit": 20, "total": 0, "total_pages": 0}}

@router.get("/attendance/export-excel")
def export_attendance_excel(
    month: int,
    year: int,
    employee_id: Optional[str] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Export attendance data from attendance_list table to Excel"""
    if current_user.role not in ["Admin", "Manager", "HR"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    try:
        # Query attendance_list with same filters as the displayed table
        query = db.query(AttendanceList).filter(
            and_(
                AttendanceList.month == str(month),
                AttendanceList.year == year
            )
        )
        
        # Apply search filter
        if search and search.strip():
            search_term = search.strip().lower()
            query = query.filter(
                or_(
                    func.lower(AttendanceList.name).contains(search_term),
                    cast(AttendanceList.empid, String).contains(search_term)
                )
            )
        
        # Apply employee_id filter
        if employee_id:
            try:
                empid_int = int(employee_id) if employee_id and str(employee_id).isdigit() else None
                if empid_int:
                    query = query.filter(AttendanceList.empid == empid_int)
            except:
                pass
        
        # Get all records (no pagination for export)
        records = query.order_by(AttendanceList.name).all()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Summary"
        
        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Title row
        month_name = datetime(year, month, 1).strftime("%B %Y")
        ws.merge_cells('A1:Q1')
        ws['A1'] = f"ATTENDANCE SUMMARY - {month_name.upper()}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 25
        
        # Header row
        headers = [
            "EMPLOYEE-NAME", "EMP-ID", "TOTAL", "WORK", "W.O", "HOLIDAYS",
            "PRESENT", "ABSENT", "HALFDAYS", "LATE", "LOPs", "CL", "SL",
            "COMP", "PAYBLE", "MONTH", "YEAR"
        ]
        row_num = 3
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_align
        
        # Data rows
        row_num = 4
        for rec in records:
            ws.cell(row=row_num, column=1, value=rec.name or '').border = border
            ws.cell(row=row_num, column=2, value=rec.empid or '').border = border
            ws.cell(row=row_num, column=2).alignment = center_align
            ws.cell(row=row_num, column=3, value=float(rec.total_days) if rec.total_days else 0).border = border
            ws.cell(row=row_num, column=3).alignment = center_align
            ws.cell(row=row_num, column=4, value=float(rec.working_days) if rec.working_days else 0).border = border
            ws.cell(row=row_num, column=4).alignment = center_align
            ws.cell(row=row_num, column=5, value=rec.week_offs or 0).border = border
            ws.cell(row=row_num, column=5).alignment = center_align
            ws.cell(row=row_num, column=6, value=float(rec.holi_days) if rec.holi_days else 0).border = border
            ws.cell(row=row_num, column=6).alignment = center_align
            ws.cell(row=row_num, column=7, value=float(rec.presents) if rec.presents else 0).border = border
            ws.cell(row=row_num, column=7).alignment = center_align
            ws.cell(row=row_num, column=8, value=float(rec.absents) if rec.absents else 0).border = border
            ws.cell(row=row_num, column=8).alignment = center_align
            ws.cell(row=row_num, column=9, value=float(rec.half_days) if rec.half_days else 0).border = border
            ws.cell(row=row_num, column=9).alignment = center_align
            ws.cell(row=row_num, column=10, value=rec.late_logs or 0).border = border
            ws.cell(row=row_num, column=10).alignment = center_align
            ws.cell(row=row_num, column=11, value=float(rec.lops) if rec.lops else 0).border = border
            ws.cell(row=row_num, column=11).alignment = center_align
            ws.cell(row=row_num, column=12, value=float(rec.cl) if rec.cl else 0).border = border
            ws.cell(row=row_num, column=12).alignment = center_align
            ws.cell(row=row_num, column=13, value=float(rec.sl) if rec.sl else 0).border = border
            ws.cell(row=row_num, column=13).alignment = center_align
            ws.cell(row=row_num, column=14, value=float(rec.comp_offs) if rec.comp_offs else 0).border = border
            ws.cell(row=row_num, column=14).alignment = center_align
            ws.cell(row=row_num, column=15, value=float(rec.payble_days) if rec.payble_days else 0).border = border
            ws.cell(row=row_num, column=15).alignment = center_align
            ws.cell(row=row_num, column=16, value=month).border = border  # Month number (10)
            ws.cell(row=row_num, column=16).alignment = center_align
            ws.cell(row=row_num, column=17, value=year).border = border  # Year (2025)
            ws.cell(row=row_num, column=17).alignment = center_align
            row_num += 1
        
        # Set column widths
        column_widths = [25, 15, 12, 12, 10, 12, 12, 12, 12, 10, 10, 10, 10, 10, 12, 10, 10]
        for idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"attendance_summary_{year}_{str(month).zfill(2)}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        print(f"Error exporting attendance to Excel: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error exporting attendance: {str(e)}")

@router.post("/attendance/generate")
def generate_attendance(
    request: AttendanceGenerateRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Generate attendance records based on attendance cycle"""
    if current_user.role not in ["Admin", "Manager", "HR"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    try:
        from calendar import monthrange
        import traceback
        
        month = request.month
        year = request.year
        employee_id = request.employee_id
        
        # Get attendance cycle (default to id=1)
        cycle = db.query(AttendanceCycle).filter(AttendanceCycle.id == 1).first()
        if not cycle:
            raise HTTPException(status_code=404, detail="Attendance cycle not found. Please create an attendance cycle first.")
        
        # Calculate cycle dates
        start_day = cycle.attendance_cycle_start_date
        end_day = cycle.attendance_cycle_end_date
        
        # Calculate first_date
        if start_day <= end_day:
            # Same month cycle (e.g., 1-31)
            first_date = date(year, month, start_day)
        else:
            # Cross month cycle (e.g., 26-25)
            if month == 1:
                prev_month = 12
                prev_year = year - 1
            else:
                prev_month = month - 1
                prev_year = year
            first_date = date(prev_year, prev_month, start_day)
        
        # Calculate last_date (always in selected month)
        last_date = date(year, month, end_day)
        
        # Get employees to process (only active employees, matching reference code Status = '1')
        if employee_id:
            employees = db.query(User).filter(
                and_(
                    User.empid == employee_id,
                    User.is_active == True
                )
            ).all()
            print(f"Found {len(employees)} active employee(s) with empid: {employee_id}")
        else:
            employees = db.query(User).filter(User.is_active == True).all()
            print(f"Processing all {len(employees)} active employees")
        
        if not employees:
            raise HTTPException(status_code=404, detail=f"No employees found for employee_id: {employee_id}")
        
        # Get all dates in cycle
        current_date = first_date
        dates_in_cycle = []
        while current_date <= last_date:
            dates_in_cycle.append(current_date)
            current_date += timedelta(days=1)
        
        total_days = len(dates_in_cycle)
        month_days = monthrange(year, month)[1]
        
        # Get all holidays in cycle range (will be filtered per employee by branch_id)
        all_holiday_records = db.query(Holiday).filter(
            and_(
                Holiday.date >= first_date,
                Holiday.date <= last_date
            )
        ).all()
        
        all_week_off_dates = set()
        all_week_off_records = db.query(WeekOffDate).filter(
            and_(
                WeekOffDate.date >= first_date,
                WeekOffDate.date <= last_date
            )
        ).all()
        for wo in all_week_off_records:
            wo_date = wo.date if isinstance(wo.date, date) else wo.date.date() if hasattr(wo.date, 'date') else wo.date
            all_week_off_dates.add(wo_date)
        
        # Process each employee
        generated_count = 0
        errors = []
        for employee in employees:
            try:
                print(f"Processing employee: {employee.name} (ID: {employee.empid})")
                
                # Calculate employee-specific date range based on DOJ and emp_inactive_date
                # Logic:
                # - If DOJ exists and is within the generate month (first_date <= DOJ <= last_date):
                #   → Start from DOJ (only count from DOJ onwards: presents, week-offs, holidays, leaves, half-days)
                # - If DOJ exists but is before first_date (DOJ < first_date):
                #   → Take full month (start from first_date)
                # - If DOJ is null/empty:
                #   → Treat as previous employee, take full month (start from first_date)
                if employee.doj:
                    # DOJ exists: empStartDate = max(DOJ, first_date)
                    # This ensures: if DOJ is in period, use DOJ; if DOJ is before period, use first_date (full month)
                    emp_start_date = max(employee.doj, first_date)
                else:
                    # DOJ is null/empty: previous employee, use full month
                    emp_start_date = first_date
                
                # Get emp_inactive_date (check if field exists using getattr for safety)
                emp_inactive_date = getattr(employee, 'emp_inactive_date', None)
                
                # If emp_inactive_date exists, end at inactive date (but not after last_date)
                # If emp_inactive_date is null/empty, use full month
                if emp_inactive_date:
                    # emp_inactive_date exists: only count up to inactive date
                    emp_end_date = min(emp_inactive_date, last_date)
                else:
                    # emp_inactive_date is null/empty: use full month
                    emp_end_date = last_date
                
                print(f"  Employee date range: DOJ={employee.doj}, emp_inactive_date={emp_inactive_date}, "
                      f"emp_start_date={emp_start_date}, emp_end_date={emp_end_date}")
                
                # Calculate employee total days in their active period
                emp_total_days = (emp_end_date - emp_start_date).days + 1
                if emp_total_days <= 0:
                    print(f"  Skipping employee {employee.name}: no days in period")
                    continue
                
                # Count week-offs and holidays within employee's effective period ONLY
                # IMPORTANT: Only count holidays/week-offs from emp_start_date to emp_end_date
                # Dates before DOJ should NOT count holidays/week-offs
                # Dates after emp_inactive_date should NOT count holidays/week-offs
                # Holidays are filtered by employee's branch_id (same logic as punch.jsx)
                emp_week_off_count = 0
                emp_holiday_count = 0
                
                # Build employee-specific holiday dates set based on branch_id
                # Logic: If employee's branch_id matches an entry in holiday's holiday_permissions → Count as holiday
                # If employee's branch_id is null or empty, default to branch_id = 1
                emp_holiday_dates = set()
                # Use employee's branch_id, or default to 1 if null/empty
                employee_branch_id = employee.branch_id if employee.branch_id else 1
                
                for holiday in all_holiday_records:
                    h_date = holiday.date if isinstance(holiday.date, date) else holiday.date.date() if hasattr(holiday.date, 'date') else holiday.date
                    # Check if holiday is within employee's effective period
                    if h_date < emp_start_date or h_date > emp_end_date:
                        continue
                    
                    # If holiday has no permissions, don't count it
                    if not holiday.holiday_permissions or len(holiday.holiday_permissions) == 0:
                        continue
                    
                    # Check if employee's branch_id (or default 1) exists in the holiday's holiday_permissions array
                    # Logic: If employee's branch_id matches any branch_id in holiday_permissions → Show holiday
                    # Example:
                    #   - Employee branch_id = 1 (or null/empty, defaulted to 1)
                    #   - holiday_permissions = [{"branch_id": 1, "branch_name": "CORPORATE OFFICE"}, {"branch_id": 2, "branch_name": "MUMBAI"}]
                    #   - Since branch_id 1 exists in the array → Count as holiday ✓
                    #   - If holiday_permissions = [{"branch_id": 2, "branch_name": "MUMBAI"}] → Don't count as holiday ✗
                    if any(perm.get('branch_id') == employee_branch_id for perm in holiday.holiday_permissions):
                        emp_holiday_dates.add(h_date)
                
                for day in [emp_start_date + timedelta(days=x) for x in range(emp_total_days)]:
                    # Check if it's a week-off (employee-specific or global)
                    is_week_off = False
                    week_off_check = db.query(WeekOffDate).filter(
                        and_(
                            or_(
                                WeekOffDate.employee_id == employee.empid,
                                WeekOffDate.employee_id == "0"
                            ),
                            WeekOffDate.date == day
                        )
                    ).first()
                    if week_off_check:
                        is_week_off = True
                        emp_week_off_count += 1
                    
                    # Check if it's a holiday (filtered by employee's branch_id)
                    if day in emp_holiday_dates:
                        emp_holiday_count += 1
                
                # Get approved leaves for this employee in cycle range (ONLY APPROVED)
                leave_records = db.query(Leave).filter(
                    and_(
                        Leave.empid == employee.empid,
                        Leave.status == 'approved',  # Only count approved leaves
                        Leave.from_date <= emp_end_date,
                        Leave.to_date >= emp_start_date
                    )
                ).all()
                
                # Create a set of leave dates to check before processing attendance
                # Priority: Leave > Attendance (if leave exists on a date, skip attendance processing)
                leave_dates_set = set()
                leave_dates_by_type = {}  # {date: leave_type} for quick lookup
                
                # Calculate leaves by type (sum of durations)
                cl = 0.0  # Casual Leave
                sl = 0.0  # Sick Leave
                comp_offs = 0.0  # Comp-Off Leave
                lop_leaves = 0.0  # LOP Leave
                
                for leave in leave_records:
                    # Get all dates in the leave range within employee period
                    leave_start = max(leave.from_date, emp_start_date)
                    leave_end = min(leave.to_date, emp_end_date)
                    
                    # Add all dates in leave range to the set
                    current_leave_date = leave_start
                    while current_leave_date <= leave_end:
                        # Skip if it's a week-off (week-off takes priority over leave)
                        is_week_off = False
                        week_off_check = db.query(WeekOffDate).filter(
                            and_(
                                or_(
                                    WeekOffDate.employee_id == employee.empid,
                                    WeekOffDate.employee_id == "0"
                                ),
                                WeekOffDate.date == current_leave_date
                            )
                        ).first()
                        if week_off_check:
                            is_week_off = True
                        
                        # Only add to leave dates if not a week-off or holiday (use employee-specific holiday dates)
                        if not is_week_off and current_leave_date not in emp_holiday_dates:
                            leave_dates_set.add(current_leave_date)
                            leave_dates_by_type[current_leave_date] = leave.leave_type
                        current_leave_date += timedelta(days=1)
                    
                    # Calculate leave duration for counting (excluding week-offs and holidays)
                    leave_duration = 0
                    temp_date = leave_start
                    while temp_date <= leave_end:
                        is_week_off = False
                        week_off_check = db.query(WeekOffDate).filter(
                            and_(
                                or_(
                                    WeekOffDate.employee_id == employee.empid,
                                    WeekOffDate.employee_id == "0"
                                ),
                                WeekOffDate.date == temp_date
                            )
                        ).first()
                        if week_off_check:
                            is_week_off = True
                        
                        if not is_week_off and temp_date not in emp_holiday_dates:
                            leave_duration += 1
                        temp_date += timedelta(days=1)
                    
                    if leave_duration <= 0:
                        continue
                    
                    leave_type = leave.leave_type.lower().strip() if leave.leave_type else ''
                    
                    # Match leave types (case-insensitive, handle variations)
                    # Check for exact matches and common variations
                    if leave_type in ['casual', 'casual leave', 'casualleave']:
                        cl += leave_duration
                    elif leave_type in ['sick', 'sick leave', 'sickleave']:
                        sl += leave_duration
                    elif leave_type in ['comp-off', 'compensatory', 'compensatory-off', 'comp off', 'compensatory off', 'compensatoryoff', 'compoff']:
                        comp_offs += leave_duration
                    elif leave_type in ['lop', 'loss of pay', 'loss-of-pay', 'lop leave', 'lossofpay', 'lopleave']:
                        lop_leaves += leave_duration
                    else:
                        # Debug: Print unmatched leave types
                        print(f"  Warning: Unmatched leave type '{leave.leave_type}' for employee {employee.name}")
                
                # Debug: Print leave counts
                print(f"  Leaves for {employee.name}: CL={cl}, SL={sl}, COMP={comp_offs}, LOP={lop_leaves}, Total Approved Leaves={len(leave_records)}, Leave Dates={len(leave_dates_set)}")
                if len(leave_records) > 0:
                    print(f"    Leave details: {[(l.leave_type, l.status, l.from_date, l.to_date, l.duration) for l in leave_records[:5]]}")
                
                # Process attendance records
                presents = 0
                half_day_count = 0
                late_log_count = 0
                processed_dates = set()
                
                # Get all punch logs for this employee in their period
                from datetime import datetime as dt_datetime
                emp_end_datetime = dt_datetime.combine(emp_end_date, dt_datetime.max.time())
                
                punch_logs_query = db.query(PunchLog).filter(
                    and_(
                        PunchLog.employee_id == str(employee.empid),
                        PunchLog.date >= emp_start_date,
                        PunchLog.date <= emp_end_date
                    )
                ).all()
                
                # Group punch logs by date
                from collections import defaultdict
                punches_by_date = defaultdict(list)
                for log in punch_logs_query:
                    log_date = log.date if isinstance(log.date, date) else log.date.date() if hasattr(log.date, 'date') else log.date
                    if emp_start_date <= log_date <= emp_end_date:
                        punches_by_date[log_date].append(log)
                
                # Process each date with punch logs
                for t_date, logs in punches_by_date.items():
                    # Priority: Week-off > Holiday > Leave > Attendance
                    # Skip if week-off or holiday (even if punch logs exist, don't count attendance)
                    is_week_off = False
                    week_off_check = db.query(WeekOffDate).filter(
                        and_(
                            or_(
                                WeekOffDate.employee_id == employee.empid,
                                WeekOffDate.employee_id == "0"
                            ),
                            WeekOffDate.date == t_date
                        )
                    ).first()
                    if week_off_check:
                        is_week_off = True
                    
                    if is_week_off:
                        # Skip attendance processing - only count as week-off (already counted in emp_week_off_count)
                        print(f"    Skipping attendance for {t_date} - is week-off (punch logs ignored)")
                        continue
                    
                    if t_date in emp_holiday_dates:
                        # Skip attendance processing - only count as holiday (already counted in emp_holiday_count)
                        print(f"    Skipping attendance for {t_date} - is holiday (punch logs ignored)")
                        continue
                    
                    # IMPORTANT: If date has an approved leave, skip attendance processing
                    # Even if punch logs exist, we should NOT count attendance, only the leave
                    if t_date in leave_dates_set:
                        # This date has a leave, skip attendance processing
                        # The leave is already counted above
                        print(f"    Skipping attendance for {t_date} - has approved leave ({leave_dates_by_type.get(t_date, 'unknown')})")
                        continue
                    
                    # Filter valid punch times (not midnight)
                    from datetime import time as dt_time
                    midnight = dt_time(0, 0, 0)
                    valid_logs = [log for log in logs if log.punch_time and log.punch_time.time() != midnight]
                    
                    if not valid_logs:
                        continue
                    
                    # Calculate in-time and out-time
                    punch_times = [log.punch_time for log in valid_logs]
                    min_time = min(punch_times)
                    max_time = max(punch_times)
                    
                    in_time = min_time.time()
                    out_time = max_time.time()
                    
                    # Calculate duration
                    duration_seconds = (max_time - min_time).total_seconds()
                    duration_hours = duration_seconds / 3600
                    
                    if duration_hours < 0:
                        continue
                    
                    # Check for late log (previous code - calculate normally)
                    late_log_time = cycle.late_log_time
                    is_late = in_time > late_log_time
                    
                    # Get duration thresholds
                    full_day_hours = (cycle.full_day_duration.hour * 3600 + cycle.full_day_duration.minute * 60) / 3600
                    half_day_hours = (cycle.half_day_duration.hour * 3600 + cycle.half_day_duration.minute * 60) / 3600
                    
                    # Classify attendance
                    if duration_hours >= full_day_hours:
                        presents += 1
                        if is_late:
                            late_log_count += 1
                    elif duration_hours >= half_day_hours:
                        half_day_count += 1
                        if is_late:
                            late_log_count += 1
                    
                    processed_dates.add(t_date)
                
                # Calculate uncovered absents (dates not processed, not weekoffs, not holidays, not leaves)
                potential_absents = 0
                for day in [emp_start_date + timedelta(days=x) for x in range(emp_total_days)]:
                    is_week_off = False
                    week_off_check = db.query(WeekOffDate).filter(
                        and_(
                            or_(
                                WeekOffDate.employee_id == employee.empid,
                                WeekOffDate.employee_id == "0"
                            ),
                            WeekOffDate.date == day
                        )
                    ).first()
                    if week_off_check:
                        is_week_off = True
                    
                    # Skip if: processed (has attendance), week-off, holiday, or has leave
                    if (day not in processed_dates and 
                        not is_week_off and 
                        day not in emp_holiday_dates and
                        day not in leave_dates_set):  # Exclude leave dates
                        potential_absents += 1
                
                # Calculate total paid leaves (except LOP)
                total_paid_leaves = cl + sl + comp_offs
                
                # Absents = potential_absents - total_paid_leaves (minimum 0)
                absents = max(0.0, float(potential_absents) - total_paid_leaves)
                
                # Calculate half days as decimal
                half_days = half_day_count * 0.5
                
                # Calculate working days
                working_days = emp_total_days - emp_week_off_count - emp_holiday_count
                
                # Check if employee has is_late flag set to True
                # If is_late is False or None, set late_log_count to 0
                # Employee is already a User object, so we can access is_late directly
                employee_is_late_enabled = getattr(employee, 'is_late', False) or False
                if not employee_is_late_enabled:
                    late_log_count = 0
                
                # Calculate late log deduction (every 3 late logs = 0.5 day)
                late_log_deduction = (late_log_count // 3) * 0.5
                
                # Calculate payable days (before scaling to month)
                # payableDays = presents + halfDays + totalPaidLeaves + weekOffs + holidays - lateDeduction
                payable_days = presents + half_days + total_paid_leaves + emp_week_off_count + emp_holiday_count - late_log_deduction
                
                # Calculate LOPs
                lops = float(emp_total_days) - payable_days
                if lops < 0:
                    lops = 0.0
                
                # Calculate final payable days (scaled to selected month)
                # Reference: finalPayableDays = (payableDays / empTotalDays) * selectedMonthDays
                # But user requirement: use full cycle range (total_days) as denominator, not emp_total_days
                # Example: Employee 1030 with DOJ 2025-12-16, period 2025-11-26 to 2025-12-25
                #   - emp_total_days = 10 (2025-12-16 to 2025-12-25)
                #   - total_days = 30 (full cycle: 2025-11-26 to 2025-12-25)
                #   - finalPayableDays = (payableDays / 30) * 31
                if total_days > 0:
                    final_payable_days = (payable_days / total_days) * month_days
                else:
                    final_payable_days = 0.0
                
                # Special rounding logic for payble_days
                # Rounding rules:
                # - 0.1-0.4 → rounds to 0.5 (e.g., 28.1, 28.2, 28.3, 28.4 → 28.5)
                # - 0.6-0.9 → rounds to 1.0 (e.g., 28.6, 28.7, 28.8, 28.9 → 29.0)
                # - 0.5 → stays as is (e.g., 28.5 → 28.5)
                # - Whole numbers → stay as is (e.g., 28.0 → 28.0)
                def round_payble_days(value):
                    # Handle edge cases
                    if value is None:
                        return 0.0
                    if value <= 0:
                        return 0.0
                    
                    # Round to 1 decimal place first to handle floating point precision issues
                    value = round(value, 1)
                    
                    # Check if it's a whole number
                    if value == int(value):
                        return float(int(value))
                    
                    integer_part = int(value)
                    # Round to 1 decimal to avoid floating point precision issues
                    decimal_part = round(value - integer_part, 1)
                    
                    # Apply rounding rules
                    if 0.1 <= decimal_part <= 0.4:
                        return float(integer_part) + 0.5
                    elif 0.6 <= decimal_part <= 0.9:
                        return float(integer_part) + 1.0
                    elif decimal_part == 0.5:
                        return value
                    else:
                        # For values outside 0.1-0.9 range (shouldn't happen), return as is
                        return value
                
                final_payable_days = round_payble_days(final_payable_days)
                
                # Debug: Print calculated values
                print(f"  Calculated for {employee.name}: emp_total_days={emp_total_days}, presents={presents}, "
                      f"absents={absents}, half_days={half_days}, week_offs={emp_week_off_count}, "
                      f"holidays={emp_holiday_count}, cl={cl}, sl={sl}, comp_offs={comp_offs}, "
                      f"late_logs={late_log_count}, late_deduction={late_log_deduction}, "
                      f"payable_days={payable_days}, final_payable_days={final_payable_days}, lops={lops}")
                
                # Check if record already exists for this employee, month, year
                # empid column is INTEGER - convert employee.empid (string) to integer if numeric
                # Otherwise use employee.id as fallback
                try:
                    empid_value = int(employee.empid) if employee.empid and str(employee.empid).isdigit() else employee.id
                except:
                    empid_value = employee.id
                
                # Check if record exists based on empid, month, and year
                # If record exists, update it; otherwise, insert new record
                existing = db.query(AttendanceList).filter(
                    and_(
                        AttendanceList.empid == empid_value,
                        AttendanceList.month == str(month),
                        AttendanceList.year == year
                    )
                ).first()
                
                attendance_data = {
                    'name': employee.name.upper() if employee.name else employee.name,  # Uppercase name like reference
                    'empid': empid_value,
                    'doj': employee.doj if employee.doj else emp_start_date,  # Use DOJ or fallback to start date
                    'from_date': emp_start_date,  # Employee-specific start date
                    'to_date': emp_end_date,  # Employee-specific end date
                    'total_days': float(month_days),  # Selected month days (not emp_total_days)
                    'working_days': float(working_days),
                    'week_offs': emp_week_off_count,
                    'holi_days': float(emp_holiday_count),
                    'presents': float(presents),
                    'absents': float(absents),
                    'half_days': float(half_days),  # Already in decimal (half_day_count * 0.5)
                    'late_logs': late_log_count,
                    'cl': float(cl),
                    'sl': float(sl),
                    'comp_offs': float(comp_offs),
                    'payble_days': final_payable_days,  # Already rounded with special logic
                    'lops': round(lops, 1),
                    'year': year,
                    'month': str(month),
                    'status': 1,  # Status = 1 (active) like reference code
                    'updated_by': current_user.name or current_user.empid,
                    'updated_date': get_ist_now()
                }
                
                if existing:
                    # Update existing record - update all fields (matching reference code)
                    for key, value in attendance_data.items():
                        setattr(existing, key, value)
                    # Note: image_base64 is fetched from users table when querying, not stored here
                    print(f"  ✓ Updated attendance for employee {employee.empid} ({employee.name}) - Month: {month}, Year: {year}")
                else:
                    # Create new record
                    # Try with string empid first, if column is still INTEGER, use integer ID
                    try:
                        new_record = AttendanceList(**attendance_data)
                        db.add(new_record)
                        db.flush()  # Flush to get the ID and check for errors
                        print(f"  ✓ Created attendance record for employee {employee.empid} ({employee.name}) - ID: {new_record.id}, Month: {month}, Year: {year}")
                    except Exception as create_error:
                        print(f"  ✗ Failed to create record for {employee.empid}: {create_error}")
                        import traceback
                        traceback.print_exc()
                        raise  # Re-raise to be caught by outer exception handler
                
                generated_count += 1
                print(f"  Successfully processed {employee.name}. Total generated so far: {generated_count}")
            except Exception as e:
                error_msg = f"Error processing employee {employee.empid} ({employee.name if employee else 'Unknown'}): {str(e)}"
                print(error_msg)
                traceback.print_exc()
                errors.append(error_msg)
                # Don't continue silently - log the error but still try to commit others
                continue
        
        if errors:
            print(f"Errors encountered: {errors}")
        
        if generated_count == 0:
            db.rollback()
            error_details = f"No attendance records were generated. "
            if errors:
                error_details += f"Errors encountered: {', '.join(errors[:3])}"  # Show first 3 errors
            else:
                error_details += "Please check if employees have punch logs or if there are any errors in the logs."
            raise HTTPException(
                status_code=400, 
                detail=error_details
            )
        
        db.commit()
        print(f"Successfully committed {generated_count} attendance record(s)")
        
        return {
            "message": f"Attendance generated successfully for {generated_count} employee(s)",
            "generated_count": generated_count,
            "month": month,
            "year": year,
            "cycle_dates": {
                "from_date": first_date.isoformat(),
                "to_date": last_date.isoformat()
            }
        }
    
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"Error generating attendance: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to generate attendance: {str(e)}")

@router.get("/attendance/history")
def get_attendance_history(
    date: Optional[str] = None,
    filter: Optional[str] = "all",
    db: Session = Depends(get_db)
):
    """Get attendance list for a specific date with min/max times from punch_logs"""
    if not date:
        date_str = datetime.now().date().isoformat()
    else:
        date_str = date
    
    try:
        target_date = datetime.fromisoformat(date_str).date()
    except:
        target_date = datetime.now().date()
    
    # Get all employees - limit to 500 for performance
    employees = db.query(User).filter(User.role.in_(['Employee', 'Manager'])).limit(500).all()
    
    # Batch load attendance records to avoid N+1 queries
    empids = [emp.empid for emp in employees if emp.empid]
    attendance_map = {}
    if empids:
        attendance_records = db.query(Attendance).filter(
            and_(
                Attendance.employee_id.in_(empids),
                Attendance.date == target_date
            )
        ).all()
        attendance_map = {att.employee_id: att for att in attendance_records}
    
    # Batch load punch logs to avoid N+1 queries
    punch_logs_map = {}
    if empids:
        punch_logs_all = db.query(PunchLog).filter(
            and_(
                PunchLog.employee_id.in_(empids),
                PunchLog.date == target_date
            )
        ).order_by(PunchLog.employee_id, PunchLog.punch_time).limit(1000).all()
        for log in punch_logs_all:
            if log.employee_id not in punch_logs_map:
                punch_logs_map[log.employee_id] = []
            punch_logs_map[log.employee_id].append(log)
    
    result = []
    for emp in employees:
        # Get attendance record from map
        attendance = attendance_map.get(emp.empid)
        
        # If attendance record exists with check_in/check_out, use those (manually modified)
        if attendance and attendance.check_in and attendance.check_out:
            in_time = attendance.check_in.strftime('%H:%M')
            out_time = attendance.check_out.strftime('%H:%M')
            min_time = attendance.check_in
            max_time = attendance.check_out
        else:
            # Otherwise, get min/max from punch_logs map
            punch_logs = punch_logs_map.get(emp.empid, [])
            
            # Get min and max times
            min_time = None
            max_time = None
            if punch_logs:
                times = [log.punch_time for log in punch_logs]
                min_time = min(times)
                max_time = max(times)
            
            # Format times
            in_time = min_time.strftime('%H:%M') if min_time else None
            out_time = max_time.strftime('%H:%M') if max_time else None
        
        # Calculate hours
        hours = None
        if min_time and max_time:
            hours = round((max_time - min_time).total_seconds() / 3600, 2)
        
        result.append({
            "id": attendance.id if attendance else None,
            "employee_id": emp.empid,
            "employee_name": emp.name,
            "image_base64": emp.image_base64,
            "date": target_date.isoformat(),
            "in_time": in_time,
            "out_time": out_time,
            "check_in": min_time.isoformat() if min_time else None,
            "check_out": max_time.isoformat() if max_time else None,
            "status": attendance.status if attendance else None,
            "hours": hours
        })
    
    # Apply filter if needed
    if filter != "all":
        result = [r for r in result if r["status"] == filter]
    
    return result

@router.get("/attendance/history-month")
def get_attendance_history_month(
    month: Optional[int] = None,
    year: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get attendance history for current month with all employees and dates from punch_logs"""
    from calendar import monthrange
    
    if current_user.role not in ["Admin", "Manager", "HR"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    now = datetime.now()
    if not month:
        month = now.month
    if not year:
        year = now.year
    
    # Get first and last date of the month
    first_date = date(year, month, 1)
    last_day = monthrange(year, month)[1]
    last_date = date(year, month, last_day)
    
    # Get all employees - for Manager, include their team members + themselves
    if current_user.role == "Manager":
        # Get employees reporting to this manager - limit to 200 for performance
        team_members = db.query(User).filter(
            User.report_to_id == current_user.empid,
            User.role.in_(['Employee', 'Manager'])
        ).limit(200).all()
        # Include manager themselves
        employees = list(team_members)
        employees.append(current_user)
    else:
        # Admin/HR see all employees (including HR, Admin, Manager, Employee roles) - limit to 500 for performance
        employees = db.query(User).filter(
            User.role.in_(['Employee', 'Manager', 'HR', 'Admin'])
        ).limit(500).all()
    
    # Get all punch logs for the month - limit to 10000 for performance
    punch_logs_all = db.query(PunchLog).filter(
        and_(
            PunchLog.date >= first_date,
            PunchLog.date <= last_date
        )
    ).order_by(PunchLog.employee_id, PunchLog.date, PunchLog.punch_time).limit(10000).all()
    
    # Group punch logs by employee_id (empid) and date
    # Store both punch_time and punch_type for proper calculation
    punch_map = {}
    for log in punch_logs_all:
        key = f"{log.employee_id}_{log.date.isoformat()}"
        if key not in punch_map:
            punch_map[key] = []
        punch_map[key].append({
            'punch_time': log.punch_time,
            'punch_type': log.punch_type
        })
    
    # Build attendance map from punch_logs
    # Logic: 
    # - 1 record: intime and outtime are the same (min and max are the same)
    # - 2 records: one is intime (punch_type='in'), another is outtime (punch_type='out')
    # - 3+ records: min time is intime, max time is outtime
    attendance_map = {}
    for emp in employees:
        current_date = first_date
        while current_date <= last_date:
            key = f"{emp.empid}_{current_date.isoformat()}"
            
            if key in punch_map and punch_map[key]:
                punches = punch_map[key]
                num_punches = len(punches)
                
                check_in_time = None
                check_out_time = None
                
                if num_punches == 1:
                    # 1 record: intime and outtime are the same
                    check_in_time = punches[0]['punch_time']
                    check_out_time = punches[0]['punch_time']
                elif num_punches == 2:
                    # 2 records: one is intime, another is outtime
                    in_punch = next((p for p in punches if p['punch_type'] == 'in'), None)
                    out_punch = next((p for p in punches if p['punch_type'] == 'out'), None)
                    
                    if in_punch and out_punch:
                        check_in_time = in_punch['punch_time']
                        check_out_time = out_punch['punch_time']
                    else:
                        # Fallback: if punch_type doesn't match, use min/max
                        sorted_punches = sorted(punches, key=lambda x: x['punch_time'])
                        check_in_time = sorted_punches[0]['punch_time']
                        check_out_time = sorted_punches[1]['punch_time']
                else:
                    # 3+ records: min time is intime, max time is outtime
                    sorted_punches = sorted(punches, key=lambda x: x['punch_time'])
                    check_in_time = sorted_punches[0]['punch_time']
                    check_out_time = sorted_punches[-1]['punch_time']
                
                # Calculate duration
                if check_in_time and check_out_time:
                    delta = check_out_time - check_in_time
                    hours = delta.total_seconds() / 3600
                else:
                    hours = 0
                
                # Calculate status based on hours
                status = None
                if hours >= 9:
                    status = 'P'  # Present
                elif hours >= 4.5:  # 4:30 to 8:59
                    status = 'H/D'  # Half Day
                elif hours > 0:
                    status = 'Abs'  # Absent (less than 4:30)
                
                attendance_map[key] = {
                    "check_in": check_in_time.strftime("%H:%M") if check_in_time else "00:00",
                    "check_out": check_out_time.strftime("%H:%M") if check_out_time else "00:00",
                    "status": status,
                    "hours": round(hours, 2) if hours > 0 else 0
                }
            else:
                # No records - show 00:00 with Abs status
                attendance_map[key] = {
                    "check_in": "00:00",
                    "check_out": "00:00",
                    "status": "Abs",
                    "hours": 0
                }
            
            current_date += timedelta(days=1)
    
    # Fetch week_off_dates, holidays, and leaves for the month
    week_off_dates = db.query(WeekOffDate).filter(
        and_(
            WeekOffDate.date >= first_date,
            WeekOffDate.date <= last_date
        )
    ).all()
    
    holidays = db.query(Holiday).filter(
        and_(
            Holiday.date >= first_date,
            Holiday.date <= last_date
        )
    ).all()
    
    leaves = db.query(Leave).filter(
        and_(
            Leave.from_date <= last_date,
            Leave.to_date >= first_date,
            Leave.status == 'approved'  # Only approved leaves
        )
    ).all()
    
    # Build lookup maps for faster access
    week_off_map = {}  # {date: [employee_ids]} where employee_id "0" means all
    for wod in week_off_dates:
        date_key = wod.date.isoformat()
        if date_key not in week_off_map:
            week_off_map[date_key] = []
        week_off_map[date_key].append(wod.employee_id)
    
    # Build employee-specific holiday map based on branch_id matching
    # Logic: If employee's branch_id (or default 1 if null) matches holiday's holiday_permissions → It's a holiday
    holiday_map = {}  # {empid: {date_key: True}} - per employee holiday dates
    for emp in employees:
        # Normalize employee branch_id to int (holiday_permissions may store as int or string)
        raw_emp_branch_id = emp.branch_id if emp.branch_id else 1  # Default to 1 if null/empty
        try:
            emp_branch_id = int(raw_emp_branch_id)
        except (ValueError, TypeError):
            emp_branch_id = 1
        
        emp_holiday_dates = set()
        for holiday in holidays:
            # If holiday has no permissions, skip it
            if not holiday.holiday_permissions or len(holiday.holiday_permissions) == 0:
                continue
            # Check if employee's branch_id exists in the holiday's holiday_permissions array
            # Normalize permission branch_id to int for reliable comparison
            match = False
            for perm in holiday.holiday_permissions:
                if perm and perm.get('branch_id') is not None:
                    try:
                        perm_branch_id = int(perm.get('branch_id'))
                        if perm_branch_id == emp_branch_id:
                            match = True
                            break
                    except (ValueError, TypeError):
                        continue
            if match:
                date_key = holiday.date.isoformat()
                emp_holiday_dates.add(date_key)
        holiday_map[emp.empid] = emp_holiday_dates
    
    leave_map = {}  # {empid: [(from_date, to_date, leave_type)]}
    for leave in leaves:
        # Ensure empid is string and trimmed for consistent comparison
        # Handle both string and integer empid values from leaves table
        # users table empid = punch_logs table employee_id = leaves table empid (all should match)
        empid = str(leave.empid).strip() if leave.empid else None
        if not empid:
            continue
        if empid not in leave_map:
            leave_map[empid] = []
        leave_map[empid].append((leave.from_date, leave.to_date, leave.leave_type))
    
    # Build result with all employees and all dates
    result = []
    for emp in employees:
        emp_data = {
            "employee_id": emp.empid,
            "employee_name": emp.name,
            "dates": {}
        }
        
        # Add data for each date in the month
        current_date = first_date
        while current_date <= last_date:
            date_str = current_date.strftime("%d-%m-%Y")
            date_key = current_date.isoformat()
            key = f"{emp.empid}_{date_key}"
            
            # Get base data from attendance_map (already calculated from punch_logs)
            base_data = attendance_map.get(key, {
                "check_in": "00:00",
                "check_out": "00:00",
                "status": "Abs",
                "hours": 0
            })
            
            # Priority order: 1. Leaves > 2. Week Off > 3. Holidays > 4. Punch Logs
            # If any of the first three match, display that status and set times to 00:00
            # Even if punch_logs exist for that date, the status should be from leaves/week_off/holidays
            
            final_status = None
            special_status_found = False  # Track if leave/week_off/holiday is found
            
            # 1. Check leaves table first (highest priority)
            # Matching logic: users table empid = punch_logs table employee_id = leaves table empid (all should match)
            # Check if current_date falls within from_date to to_date (inclusive)
            emp_empid_str = str(emp.empid).strip() if emp.empid else None
            if emp_empid_str and emp_empid_str in leave_map:
                for from_date, to_date, leave_type in leave_map[emp_empid_str]:
                    # Check if current_date is within the leave period (inclusive: from_date <= current_date <= to_date)
                    if from_date <= current_date <= to_date:
                        final_status = leave_type.upper()  # e.g., "SICK", "CASUAL", "ANNUAL", "EMERGENCY", "OTHER"
                        special_status_found = True
                        break  # Found matching leave, no need to check other leaves
            
            # 2. Check week_off_dates (second priority)
            if not special_status_found:
                if date_key in week_off_map:
                    # Check if this employee is affected (employee_id "0" means all, or specific empid)
                    week_off_employee_ids = week_off_map[date_key]
                    if "0" in week_off_employee_ids or str(emp.empid).strip() in [str(eid).strip() for eid in week_off_employee_ids]:
                        final_status = "WO"
                        special_status_found = True
            
            # 3. Check holidays (third priority) - filtered by employee's branch_id
            if not special_status_found:
                # Get employee-specific holiday dates (filtered by branch_id)
                emp_holiday_dates = holiday_map.get(emp.empid, set())
                if date_key in emp_holiday_dates:
                    final_status = "Holiday"
                    special_status_found = True
            
            # 4. If no special status found, use punch_logs status (fourth priority)
            if not special_status_found:
                final_status = base_data["status"]  # Use status from punch_logs (P, H/D, Abs)
            
            # If special status found (leave/week_off/holiday):
            # - For leaves: set times to 00:00 (as before)
            # - For week_off/holiday: keep the times from punch_logs if they exist, only change status
            if special_status_found:
                # Check if this is a leave (not week off or holiday)
                is_leave = False
                emp_empid_str = str(emp.empid).strip() if emp.empid else None
                if emp_empid_str and emp_empid_str in leave_map:
                    for from_date, to_date, leave_type in leave_map[emp_empid_str]:
                        if from_date <= current_date <= to_date:
                            is_leave = True
                            break
                
                # Only set times to 00:00 for leaves, not for week off/holiday
                if is_leave:
                    base_data["check_in"] = "00:00"
                    base_data["check_out"] = "00:00"
                    base_data["hours"] = 0
                # For week off/holiday, keep the times from punch_logs (already in base_data)
            
            # If no status at all, set to Abs
            if final_status is None:
                final_status = "Abs"
                base_data["check_in"] = "00:00"
                base_data["check_out"] = "00:00"
                base_data["hours"] = 0
            
            # If no status and no records, set to Abs
            if final_status is None and base_data["status"] is None:
                final_status = "Abs"
            
            # Update the data with final status
            emp_data["dates"][date_str] = {
                "check_in": base_data["check_in"],
                "check_out": base_data["check_out"],
                "status": final_status,
                "hours": base_data["hours"]
            }
            
            current_date += timedelta(days=1)
        
        result.append(emp_data)
    
    # Also return date list for frontend - format as dd-mm-yyyy
    date_list = []
    current_date = first_date
    while current_date <= last_date:
        date_list.append(current_date.strftime("%d-%m-%Y"))
        current_date += timedelta(days=1)
    
    return {
        "employees": result,
        "dates": date_list,
        "month": month,
        "year": year
    }

@router.get("/attendance/history-month-self")
def get_attendance_history_month_self(
    month: Optional[int] = None,
    year: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get attendance history for current user only (Employee role)"""
    from calendar import monthrange
    
    now = datetime.now()
    if not month:
        month = now.month
    if not year:
        year = now.year
    
    # Get first and last date of the month
    first_date = date(year, month, 1)
    last_day = monthrange(year, month)[1]
    last_date = date(year, month, last_day)
    
    # Get only current user
    employees = [current_user]
    
    # Get all punch logs for the month for this user
    punch_logs_all = db.query(PunchLog).filter(
        and_(
            PunchLog.employee_id == current_user.empid,
            PunchLog.date >= first_date,
            PunchLog.date <= last_date
        )
    ).order_by(PunchLog.date, PunchLog.punch_time).all()
    
    # Group punch logs by date - store both punch_time and punch_type
    punch_map = {}
    for log in punch_logs_all:
        key = f"{log.employee_id}_{log.date.isoformat()}"
        if key not in punch_map:
            punch_map[key] = []
        punch_map[key].append({
            'punch_time': log.punch_time,
            'punch_type': log.punch_type
        })
    
    # Build attendance map from punch_logs
    # Logic: 
    # - 1 record: intime and outtime are the same (min and max are the same)
    # - 2 records: one is intime (punch_type='in'), another is outtime (punch_type='out')
    # - 3+ records: min time is intime, max time is outtime
    attendance_map = {}
    emp = current_user
    current_date = first_date
    while current_date <= last_date:
        key = f"{emp.empid}_{current_date.isoformat()}"
        
        if key in punch_map and punch_map[key]:
            punches = punch_map[key]
            num_punches = len(punches)
            
            check_in_time = None
            check_out_time = None
            
            if num_punches == 1:
                # 1 record: intime and outtime are the same
                check_in_time = punches[0]['punch_time']
                check_out_time = punches[0]['punch_time']
            elif num_punches == 2:
                # 2 records: one is intime, another is outtime
                in_punch = next((p for p in punches if p['punch_type'] == 'in'), None)
                out_punch = next((p for p in punches if p['punch_type'] == 'out'), None)
                
                if in_punch and out_punch:
                    check_in_time = in_punch['punch_time']
                    check_out_time = out_punch['punch_time']
                else:
                    # Fallback: if punch_type doesn't match, use min/max
                    sorted_punches = sorted(punches, key=lambda x: x['punch_time'])
                    check_in_time = sorted_punches[0]['punch_time']
                    check_out_time = sorted_punches[1]['punch_time']
            else:
                # 3+ records: min time is intime, max time is outtime
                sorted_punches = sorted(punches, key=lambda x: x['punch_time'])
                check_in_time = sorted_punches[0]['punch_time']
                check_out_time = sorted_punches[-1]['punch_time']
            
            # Calculate duration
            if check_in_time and check_out_time:
                delta = check_out_time - check_in_time
                hours = delta.total_seconds() / 3600
            else:
                hours = 0
            
            # Calculate status based on hours
            status = None
            if hours >= 9:
                status = 'P'
            elif hours >= 4.5:
                status = 'H/D'
            elif hours > 0:
                status = 'Abs'
            
            attendance_map[key] = {
                "check_in": check_in_time.strftime("%H:%M") if check_in_time else "00:00",
                "check_out": check_out_time.strftime("%H:%M") if check_out_time else "00:00",
                "status": status,
                "hours": round(hours, 2) if hours > 0 else 0
            }
        else:
            attendance_map[key] = {
                "check_in": "00:00",
                "check_out": "00:00",
                "status": "Abs",
                "hours": 0
            }
        
        current_date += timedelta(days=1)
    
    # Fetch week_off_dates, holidays, and leaves for the month
    week_off_dates = db.query(WeekOffDate).filter(
        and_(
            WeekOffDate.date >= first_date,
            WeekOffDate.date <= last_date
        )
    ).all()
    
    holidays = db.query(Holiday).filter(
        and_(
            Holiday.date >= first_date,
            Holiday.date <= last_date
        )
    ).all()
    
    leaves = db.query(Leave).filter(
        and_(
            Leave.empid == current_user.empid,
            Leave.from_date <= last_date,
            Leave.to_date >= first_date,
            Leave.status == 'approved'  # Only approved leaves
        )
    ).all()
    
    # Build lookup maps for faster access
    week_off_map = {}  # {date: [employee_ids]} where employee_id "0" means all
    for wod in week_off_dates:
        date_key = wod.date.isoformat()
        if date_key not in week_off_map:
            week_off_map[date_key] = []
        week_off_map[date_key].append(wod.employee_id)
    
    # Build employee-specific holiday map based on branch_id matching
    # Logic: If employee's branch_id (or default 1 if null) matches holiday's holiday_permissions → It's a holiday
    emp_branch_id = current_user.branch_id if current_user.branch_id else 1  # Default to 1 if null/empty
    holiday_dates = set()
    for holiday in holidays:
        # If holiday has no permissions, skip it
        if not holiday.holiday_permissions or len(holiday.holiday_permissions) == 0:
            continue
        # Check if employee's branch_id exists in the holiday's holiday_permissions array
        if any(perm.get('branch_id') == emp_branch_id for perm in holiday.holiday_permissions):
            date_key = holiday.date.isoformat()
            holiday_dates.add(date_key)
    
    leave_map = []  # [(from_date, to_date, leave_type)]
    for leave in leaves:
        leave_map.append((leave.from_date, leave.to_date, leave.leave_type))
    
    # Build result
    emp_data = {
        "employee_id": emp.empid,
        "employee_name": emp.name,
        "dates": {}
    }
    
    current_date = first_date
    while current_date <= last_date:
        date_str = current_date.strftime("%d-%m-%Y")
        date_key = current_date.isoformat()
        key = f"{emp.empid}_{date_key}"
        
        # Get base data from attendance_map
        base_data = attendance_map.get(key, {
            "check_in": "00:00",
            "check_out": "00:00",
            "status": "Abs",
            "hours": 0
        })
        
        # Priority order: 1. Leaves > 2. Week Off > 3. Holidays > 4. Punch Logs
        # If any of the first three match, display that status and set times to 00:00
        # Even if punch_logs exist for that date, the status should be from leaves/week_off/holidays
        
        final_status = None
        special_status_found = False  # Track if leave/week_off/holiday is found
        
        # 1. Check leaves table first (highest priority)
        # Check if current_date falls within from_date to to_date (inclusive)
        for from_date, to_date, leave_type in leave_map:
            if from_date <= current_date <= to_date:
                final_status = leave_type.upper()  # e.g., "SICK", "CASUAL", "ANNUAL", "EMERGENCY", "OTHER"
                special_status_found = True
                break  # Found matching leave, no need to check other leaves
        
        # 2. Check week_off_dates (second priority)
        if not special_status_found:
            if date_key in week_off_map:
                # Check if this employee is affected (employee_id "0" means all, or specific empid)
                week_off_employee_ids = week_off_map[date_key]
                if "0" in week_off_employee_ids or str(emp.empid).strip() in [str(eid).strip() for eid in week_off_employee_ids]:
                    final_status = "WO"
                    special_status_found = True
        
        # 3. Check holidays (third priority)
        if not special_status_found:
            if date_key in holiday_dates:
                final_status = "Holiday"
                special_status_found = True
        
        # 4. If no special status found, use punch_logs status (fourth priority)
        if not special_status_found:
            final_status = base_data["status"]  # Use status from punch_logs (P, H/D, Abs)
        
        # If special status found:
        # - For leaves: set times to 00:00
        # - For week off/holiday: keep the times from punch_logs if they exist
        if special_status_found:
            # Check if this is a leave (not week off or holiday)
            is_leave = False
            for from_date, to_date, leave_type in leave_map:
                if from_date <= current_date <= to_date:
                    is_leave = True
                    break
            
            # Only set times to 00:00 for leaves, not for week off/holiday
            if is_leave:
                base_data["check_in"] = "00:00"
                base_data["check_out"] = "00:00"
                base_data["hours"] = 0
            # For week off/holiday, keep the times from punch_logs (already in base_data)
        
        # If no status at all, set to Abs
        if final_status is None:
            final_status = "Abs"
            base_data["check_in"] = "00:00"
            base_data["check_out"] = "00:00"
            base_data["hours"] = 0
        
        # Update the data with final status
        emp_data["dates"][date_str] = {
            "check_in": base_data["check_in"],
            "check_out": base_data["check_out"],
            "status": final_status,
            "hours": base_data["hours"]
        }
        
        current_date += timedelta(days=1)
    
    # Build date list
    date_list = []
    current_date = first_date
    while current_date <= last_date:
        date_list.append(current_date.strftime("%d-%m-%Y"))
        current_date += timedelta(days=1)
    
    return {
        "employees": [emp_data],
        "dates": date_list,
        "month": month,
        "year": year
    }

@router.get("/attendance/today")
def get_today_attendance(
    date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get today's attendance for current user"""
    if not date:
        date_str = datetime.now().date().isoformat()
    else:
        date_str = date
    
    try:
        target_date = datetime.fromisoformat(date_str).date()
    except:
        target_date = datetime.now().date()
    
    attendance = db.query(Attendance).filter(
        and_(
            Attendance.employee_id == current_user.empid,
            Attendance.date == target_date
        )
    ).first()
    
    if not attendance:
        return None
    
    hours = 0
    if attendance.check_in and attendance.check_out:
        delta = attendance.check_out - attendance.check_in
        hours = delta.total_seconds() / 3600
    
    return {
        "id": attendance.id,
        "employee_id": attendance.employee_id,
        "date": attendance.date.isoformat(),
        "check_in": attendance.check_in.isoformat() if attendance.check_in else None,
        "check_out": attendance.check_out.isoformat() if attendance.check_out else None,
        "status": attendance.status,
        "hours": round(hours, 2) if hours > 0 else None,
        "remarks": attendance.remarks
    }

@router.get("/attendance/punch-history")
def get_punch_history(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get punch history for current user"""
    attendance_records = db.query(Attendance).filter(
        Attendance.employee_id == current_user.empid
    ).order_by(Attendance.date.desc()).limit(30).all()
    
    result = []
    for record in attendance_records:
        hours = 0
        if record.check_in and record.check_out:
            delta = record.check_out - record.check_in
            hours = delta.total_seconds() / 3600
        
        result.append({
            "id": record.id,
            "date": record.date.isoformat(),
            "check_in": record.check_in.isoformat() if record.check_in else None,
            "check_out": record.check_out.isoformat() if record.check_out else None,
            "hours": round(hours, 2) if hours > 0 else None,
            "status": record.status
        })
    
    return result

class PunchRequest(BaseModel):
    image: Optional[str] = None
    location: Optional[str] = None
    punch_description: Optional[str] = None

class ModifyAttendanceRequest(BaseModel):
    employee_id: str
    date: str
    check_in: Optional[str] = None
    check_out: Optional[str] = None
    status: str = "present"
    remarks: Optional[str] = None

@router.post("/attendance/punch-in")
def punch_in(
    punch_data: PunchRequest = Body(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Punch in for today with image - creates punch log with punch_time only"""
    today = datetime.now().date()
    current_time = datetime.now()
    
    # Create punch log entry (punch_type is required by DB but we ignore it in calculations)
    punch_log = PunchLog(
        employee_id=current_user.empid,
        employee_name=current_user.name,
        date=today,
        punch_type='punch',  # Default value, not used in calculations
        punch_time=current_time,
        image=punch_data.image,
        location=punch_data.location,
        punch_description=punch_data.punch_description,
        status='present'
    )
    db.add(punch_log)
    
    # Get all punch logs for today to determine min/max
    today_punches = db.query(PunchLog).filter(
        and_(
            PunchLog.employee_id == current_user.empid,
            PunchLog.date == today
        )
    ).order_by(PunchLog.punch_time).all()
    
    if today_punches:
        # Get min (intime) and max (outtime) from all punch times
        times = [p.punch_time for p in today_punches]
        min_time = min(times)
        max_time = max(times)
        
        # Update or create attendance record
        attendance = db.query(Attendance).filter(
            and_(
                Attendance.employee_id == current_user.empid,
                Attendance.date == today
            )
        ).first()
        
        if attendance:
            # Update check_in if this is the first punch (min time)
            if not attendance.check_in or current_time == min_time:
                attendance.check_in = min_time
                attendance.image = punch_data.image
            attendance.status = 'present'
        else:
            attendance = Attendance(
                employee_id=current_user.empid,
                employee_name=current_user.name,
                date=today,
                check_in=min_time,
                status='present',
                image=punch_data.image
            )
            db.add(attendance)
    
    db.commit()
    return {"message": "Punched in successfully"}

@router.post("/attendance/punch-out")
def punch_out(
    punch_data: PunchRequest = Body(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Punch out for today with image - creates punch log with punch_time only"""
    today = datetime.now().date()
    current_time = datetime.now()
    
    # Create punch log entry (punch_type is required by DB but we ignore it in calculations)
    punch_log = PunchLog(
        employee_id=current_user.empid,
        employee_name=current_user.name,
        date=today,
        punch_type='punch',  # Default value, not used in calculations
        punch_time=current_time,
        image=punch_data.image,
        location=punch_data.location,
        punch_description=punch_data.punch_description,
        status='present'
    )
    db.add(punch_log)
    
    # Get all punch logs for today (including the one we just added)
    today_punches = db.query(PunchLog).filter(
        and_(
            PunchLog.employee_id == current_user.empid,
            PunchLog.date == today
        )
    ).order_by(PunchLog.punch_time).all()
    
    if not today_punches:
        raise HTTPException(status_code=400, detail="No punch records found")
    
    # Calculate min (intime) and max (outtime) from all punch times
    times = [p.punch_time for p in today_punches]
    min_time = min(times)
    max_time = max(times)
    
    # Calculate hours: max_time - min_time
    total_seconds = (max_time - min_time).total_seconds()
    hours = total_seconds / 3600
    
    # Update or create attendance record
    attendance = db.query(Attendance).filter(
        and_(
            Attendance.employee_id == current_user.empid,
            Attendance.date == today
        )
    ).first()
    
    if attendance:
        attendance.check_in = min_time
        attendance.check_out = max_time
        attendance.hours = hours
        attendance.status = 'present'
    else:
        attendance = Attendance(
            employee_id=current_user.empid,
            employee_name=current_user.name,
            date=today,
            check_in=min_time,
            check_out=max_time,
            status='present',
            hours=hours,
            image=punch_data.image
        )
        db.add(attendance)
    
    db.commit()
    return {"message": "Punched out successfully"}

@router.get("/attendance/today-punches")
def get_today_punches(
    date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get all punches for today - uses min/max punch_time (ignores punch_type)"""
    if not date:
        today = datetime.now().date()
    else:
        try:
            today = datetime.fromisoformat(date).date()
        except:
            today = datetime.now().date()
    
    # Get all punch logs for today
    punch_logs = db.query(PunchLog).filter(
        and_(
            PunchLog.employee_id == current_user.empid,
            PunchLog.date == today
        )
    ).order_by(PunchLog.punch_time).all()
    
    if not punch_logs:
        return []
    
    # Calculate min (intime) and max (outtime) from all punch times
    times = [log.punch_time for log in punch_logs]
    min_time = min(times)
    max_time = max(times)
    
    # Find images for min and max times
    min_log = next((log for log in punch_logs if log.punch_time == min_time), None)
    max_log = next((log for log in punch_logs if log.punch_time == max_time), None)
    
    # Calculate duration
    delta = max_time - min_time
    hours = delta.total_seconds() / 3600
    
    # Return single punch record with min as check_in and max as check_out
    return [{
        'id': min_log.id if min_log else (max_log.id if max_log else None),
        'check_in': min_time,
        'check_out': max_time if len(punch_logs) > 1 else None,
        'check_in_image': min_log.image if min_log else None,
        'check_out_image': max_log.image if max_log and len(punch_logs) > 1 else None,
        'hours': hours if len(punch_logs) > 1 else 0,
        'status': punch_logs[0].status if punch_logs else 'present',
        'image': min_log.image if min_log else (max_log.image if max_log else None)
    }]

@router.get("/attendance/punch-calendar")
def get_punch_calendar(
    month: int,
    year: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get punch calendar data for a month - calculates hours from punch logs (max - min time)"""
    # Get all punch logs for the month
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year + 1, 1, 1)
    else:
        end_date = date(year, month + 1, 1)
    
    punch_logs = db.query(PunchLog).filter(
        and_(
            PunchLog.employee_id == current_user.empid,
            PunchLog.date >= start_date,
            PunchLog.date < end_date
        )
    ).order_by(PunchLog.date, PunchLog.punch_time).all()
    
    # Get week off dates (for current user or all employees)
    wo_dates = {}
    try:
        week_off_dates = db.query(WeekOffDate).filter(
            and_(
                or_(
                    WeekOffDate.employee_id == current_user.empid,
                    WeekOffDate.employee_id == "0"
                ),
                WeekOffDate.date >= start_date,
                WeekOffDate.date < end_date
            )
        ).all()
        for wo in week_off_dates:
            date_str = wo.date.isoformat()
            wo_dates[date_str] = "Week-Off"
    except Exception as e:
        # If table doesn't exist, continue without week offs
        print(f"Error fetching week off dates: {e}")
        wo_dates = {}
    
    # Get holidays for the month - filter by branch_id for Employee/Manager
    holidays = {}
    try:
        holiday_query = db.query(Holiday).filter(
            and_(
                Holiday.date >= start_date,
                Holiday.date < end_date
            )
        )
        
        # For ALL roles (including Admin/HR): filter by branch_id for punch calendar
        # This ensures users only see holidays relevant to their branch on the punch page
        # If user's branch_id is null or empty, default to branch_id = 1
        user_branch_id = current_user.branch_id if current_user.branch_id else 1
        
        # Filter holidays where user's branch_id (or default 1) matches an entry in holiday_permissions
        # Example: user.branch_id = 1 (or null/empty, defaulted to 1), holiday_permissions = [{"branch_id": 1, ...}] → Show holiday
        # Example: user.branch_id = 1 (or null/empty, defaulted to 1), holiday_permissions = [{"branch_id": 2, ...}] → Don't show holiday
        holiday_records = []
        all_holidays = holiday_query.all()
        for holiday in all_holidays:
            # If holiday has no permissions, don't show it
            if not holiday.holiday_permissions or len(holiday.holiday_permissions) == 0:
                continue
            # Check if user's branch_id (or default 1) exists in the holiday's holiday_permissions array
            # Logic: If logged user's branch_id matches any branch_id in holiday_permissions → Show holiday
            # Example:
            #   - User branch_id = 1 (or null/empty, defaulted to 1)
            #   - holiday_permissions = [{"branch_id": 1, "branch_name": "CORPORATE OFFICE"}, {"branch_id": 2, "branch_name": "MUMBAI"}]
            #   - Since branch_id 1 exists in the array → Show holiday ✓
            #   - If holiday_permissions = [{"branch_id": 2, "branch_name": "MUMBAI"}] → Don't show holiday ✗
            if any(perm.get('branch_id') == user_branch_id for perm in holiday.holiday_permissions):
                holiday_records.append(holiday)
        
        for holiday in holiday_records:
            date_str = holiday.date.isoformat()
            holidays[date_str] = holiday.name
    except Exception as e:
        print(f"Error fetching holidays: {e}")
        holidays = {}
    
    # Get leaves for current user (approved leaves only)
    leave_dates = {}
    try:
        leave_records = db.query(Leave).filter(
            and_(
                Leave.empid == current_user.empid,
                Leave.status == 'approved',
                or_(
                    and_(Leave.from_date <= end_date, Leave.to_date >= start_date)
                )
            )
        ).all()
        
        for leave in leave_records:
            # Generate all dates in the leave range
            current_leave_date = max(leave.from_date, start_date)
            leave_end = min(leave.to_date, end_date - timedelta(days=1))
            
            while current_leave_date <= leave_end:
                date_str = current_leave_date.isoformat()
                # Only set if not already a week-off (week-off takes priority)
                if date_str not in wo_dates:
                    leave_dates[date_str] = leave.leave_type
                current_leave_date += timedelta(days=1)
    except Exception as e:
        print(f"Error fetching leaves: {e}")
        leave_dates = {}
    
    # Group by date and calculate hours
    date_groups = {}
    for log in punch_logs:
        date_str = log.date.isoformat()
        if date_str not in date_groups:
            date_groups[date_str] = []
        date_groups[date_str].append(log.punch_time)
    
    calendar_data = []
    for date_str, times in date_groups.items():
        if len(times) >= 1:
            # Calculate hours: max(time) - min(time)
            min_time = min(times)
            max_time = max(times)
            delta = max_time - min_time
            hours = delta.total_seconds() / 3600
            
            # Format times for display
            min_time_str = min_time.strftime('%H:%M')
            max_time_str = max_time.strftime('%H:%M')
        else:
            hours = 0
            min_time_str = None
            max_time_str = None
        
        # Determine status (priority: Week-Off > Leave > Holiday > Attendance)
        if date_str in wo_dates:
            status = 'WO'
        elif date_str in leave_dates:
            status = 'Leave'
        elif date_str in holidays:
            status = 'Holiday'
        elif hours >= 9:
            status = 'P'
        elif hours >= 4.5:
            status = 'H/D'
        else:
            status = 'Abs'
        
        day_data = {
            'date': date_str,
            'hours': round(hours, 2),
            'status': status,
            'min_time': min_time_str,
            'max_time': max_time_str
        }
        
        # Add additional info
        if date_str in wo_dates:
            day_data['week_off'] = wo_dates[date_str]
        if date_str in holidays:
            day_data['holiday'] = holidays[date_str]
        if date_str in leave_dates:
            day_data['leave_type'] = leave_dates[date_str]
        
        calendar_data.append(day_data)
    
    # Add dates with no punches
    current_date = start_date
    while current_date < end_date:
        date_str = current_date.isoformat()
        if date_str not in date_groups:
            # Determine status (priority: Week-Off > Leave > Holiday > Absent)
            if date_str in wo_dates:
                status = 'WO'
            elif date_str in leave_dates:
                status = 'Leave'
            elif date_str in holidays:
                status = 'Holiday'
            else:
                status = 'Abs'
            
            day_data = {
                'date': date_str,
                'hours': 0,
                'status': status,
                'min_time': None,
                'max_time': None
            }
            
            # Add additional info
            if date_str in wo_dates:
                day_data['week_off'] = wo_dates[date_str]
            if date_str in holidays:
                day_data['holiday'] = holidays[date_str]
            if date_str in leave_dates:
                day_data['leave_type'] = leave_dates[date_str]
            
            calendar_data.append(day_data)
        current_date += timedelta(days=1)
    
    return calendar_data

@router.get("/attendance/cycle")
def get_attendance_cycle(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get attendance cycle (default id=1) - All roles can view"""
    
    try:
        cycle = db.query(AttendanceCycle).filter(AttendanceCycle.id == 1).first()
        if not cycle:
            raise HTTPException(status_code=404, detail="Attendance cycle not found")
        
        return {
            "id": cycle.id,
            "name": cycle.name,
            "shift_start_time": cycle.shift_start_time.strftime('%H:%M') if cycle.shift_start_time else None,
            "shift_end_time": cycle.shift_end_time.strftime('%H:%M') if cycle.shift_end_time else None,
            "late_log_time": cycle.late_log_time.strftime('%H:%M') if cycle.late_log_time else None,
            "full_day_duration": cycle.full_day_duration.strftime('%H:%M') if cycle.full_day_duration else None,
            "half_day_duration": cycle.half_day_duration.strftime('%H:%M') if cycle.half_day_duration else None,
            "attendance_cycle_start_date": cycle.attendance_cycle_start_date,
            "attendance_cycle_end_date": cycle.attendance_cycle_end_date,
            "birthdays_send": cycle.birthdays_send if cycle.birthdays_send else {"day": "", "time": ""},
            "anniversaries_send": cycle.anniversaries_send if cycle.anniversaries_send else {"day": "", "time": ""},
            "weekly_attendance_send": cycle.weekly_attendance_send if cycle.weekly_attendance_send else {"day": "", "time": ""},
            "created_at": cycle.created_at.isoformat() if cycle.created_at else None
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error fetching attendance cycle: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to fetch attendance cycle: {str(e)}")

@router.post("/attendance/cycle")
def create_attendance_cycle(
    cycle_data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Create attendance cycle"""
    if current_user.role not in ["Admin", "HR"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    try:
        # Check if cycle already exists
        existing = db.query(AttendanceCycle).filter(AttendanceCycle.id == 1).first()
        if existing:
            raise HTTPException(status_code=400, detail="Attendance cycle already exists. Use PUT to update.")
        
        # Parse time strings to time objects
        from datetime import time as dt_time
        
        # Parse JSONB notification settings
        birthdays_send = None
        if cycle_data.get('birthdays_send'):
            birthdays_send = cycle_data['birthdays_send']
        
        anniversaries_send = None
        if cycle_data.get('anniversaries_send'):
            anniversaries_send = cycle_data['anniversaries_send']
        
        weekly_attendance_send = None
        if cycle_data.get('weekly_attendance_send'):
            weekly_attendance_send = cycle_data['weekly_attendance_send']
        
        new_cycle = AttendanceCycle(
            id=1,  # Always use id=1
            name=cycle_data.get('name', 'Default Cycle'),
            shift_start_time=dt_time.fromisoformat(cycle_data.get('shift_start_time', '09:00')),
            shift_end_time=dt_time.fromisoformat(cycle_data.get('shift_end_time', '18:00')),
            late_log_time=dt_time.fromisoformat(cycle_data.get('late_log_time', '09:45')),
            full_day_duration=dt_time.fromisoformat(cycle_data.get('full_day_duration', '09:00')),
            half_day_duration=dt_time.fromisoformat(cycle_data.get('half_day_duration', '04:30')),
            attendance_cycle_start_date=cycle_data.get('attendance_cycle_start_date', 26),
            attendance_cycle_end_date=cycle_data.get('attendance_cycle_end_date', 25),
            birthdays_send=birthdays_send,
            anniversaries_send=anniversaries_send,
            weekly_attendance_send=weekly_attendance_send
        )
        
        db.add(new_cycle)
        db.commit()
        db.refresh(new_cycle)
        
        return {
            "message": "Attendance cycle created successfully",
            "id": new_cycle.id
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"Error creating attendance cycle: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to create attendance cycle: {str(e)}")

@router.put("/attendance/cycle")
def update_attendance_cycle(
    cycle_data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Update attendance cycle"""
    if current_user.role not in ["Admin", "HR"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    try:
        cycle = db.query(AttendanceCycle).filter(AttendanceCycle.id == 1).first()
        if not cycle:
            raise HTTPException(status_code=404, detail="Attendance cycle not found. Use POST to create.")
        
        # Parse time strings to time objects
        from datetime import time as dt_time
        
        cycle.name = cycle_data.get('name', cycle.name)
        if 'shift_start_time' in cycle_data:
            cycle.shift_start_time = dt_time.fromisoformat(cycle_data['shift_start_time'])
        if 'shift_end_time' in cycle_data:
            cycle.shift_end_time = dt_time.fromisoformat(cycle_data['shift_end_time'])
        if 'late_log_time' in cycle_data:
            cycle.late_log_time = dt_time.fromisoformat(cycle_data['late_log_time'])
        if 'full_day_duration' in cycle_data:
            cycle.full_day_duration = dt_time.fromisoformat(cycle_data['full_day_duration'])
        if 'half_day_duration' in cycle_data:
            cycle.half_day_duration = dt_time.fromisoformat(cycle_data['half_day_duration'])
        if 'attendance_cycle_start_date' in cycle_data:
            cycle.attendance_cycle_start_date = cycle_data['attendance_cycle_start_date']
        if 'attendance_cycle_end_date' in cycle_data:
            cycle.attendance_cycle_end_date = cycle_data['attendance_cycle_end_date']
        # Notification settings (JSONB)
        if 'birthdays_send' in cycle_data:
            cycle.birthdays_send = cycle_data['birthdays_send'] if cycle_data.get('birthdays_send') else None
        if 'anniversaries_send' in cycle_data:
            cycle.anniversaries_send = cycle_data['anniversaries_send'] if cycle_data.get('anniversaries_send') else None
        if 'weekly_attendance_send' in cycle_data:
            cycle.weekly_attendance_send = cycle_data['weekly_attendance_send'] if cycle_data.get('weekly_attendance_send') else None
        
        db.commit()
        db.refresh(cycle)
        
        return {
            "message": "Attendance cycle updated successfully",
            "id": cycle.id
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"Error updating attendance cycle: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to update attendance cycle: {str(e)}")

@router.get("/config/google-maps-key")
def get_google_maps_key(
    current_user: User = Depends(get_current_user)
):
    """Get Google Maps API key for frontend"""
    return {"api_key": settings.GOOGLE_MAPS_API_KEY}

@router.get("/attendance/geocode")
def reverse_geocode(
    latitude: float,
    longitude: float,
    current_user: User = Depends(get_current_user)
):
    """Reverse geocode coordinates to address using Google Maps API (proxy to avoid CSP issues)"""
    import requests
    
    if not settings.GOOGLE_MAPS_API_KEY:
        raise HTTPException(status_code=500, detail="Google Maps API key not configured")
    
    try:
        url = f"https://maps.googleapis.com/maps/api/geocode/json?latlng={latitude},{longitude}&key={settings.GOOGLE_MAPS_API_KEY}"
        response = requests.get(url, timeout=5)
        response.raise_for_status()
        data = response.json()
        
        if data.get('status') == 'OK' and data.get('results') and len(data['results']) > 0:
            return {
                "address": data['results'][0]['formatted_address'],
                "status": "OK"
            }
        else:
            # Return coordinates as fallback
            return {
                "address": f"{latitude}, {longitude}",
                "status": data.get('status', 'UNKNOWN')
            }
    except Exception as e:
        # Return coordinates as fallback on error
        return {
            "address": f"{latitude}, {longitude}",
            "status": "ERROR",
            "error": str(e)
        }

@router.get("/attendance/previous")
def get_previous_attendance(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get previous attendance records for modify page"""
    if current_user.role not in ["Admin", "Manager", "HR"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    try:
        if not start_date:
            start = datetime.now().date() - timedelta(days=30)
        else:
            start = datetime.fromisoformat(start_date).date()
        
        if not end_date:
            end = datetime.now().date()
        else:
            end = datetime.fromisoformat(end_date).date()
        
        records = db.query(Attendance).filter(
            and_(
                Attendance.date >= start,
                Attendance.date <= end
            )
        ).order_by(Attendance.date.desc(), Attendance.employee_id).all()
        
        result = []
        for record in records:
            # Get min check_in and max check_out from punch logs
            check_in_time = None
            check_out_time = None
            hours = 0
            
            try:
                punch_logs = db.query(PunchLog).filter(
                    and_(
                        PunchLog.employee_id == record.employee_id,
                        PunchLog.date == record.date
                    )
                ).order_by(PunchLog.punch_time).all()
                
                if punch_logs:
                    # Use min time as intime and max time as outtime (ignore punch_type)
                    times = [p.punch_time for p in punch_logs]
                    check_in_time = min(times)
                    if len(times) > 1:
                        check_out_time = max(times)
                    else:
                        # If only 1 record, both intime and outtime are the same
                        check_out_time = check_in_time
            except:
                pass
            
            # Fallback to attendance record if no punch logs
            if not check_in_time and record.check_in:
                check_in_time = record.check_in
            if not check_out_time and record.check_out:
                check_out_time = record.check_out
            
            # Calculate hours
            if check_in_time and check_out_time:
                delta = check_out_time - check_in_time
                hours = delta.total_seconds() / 3600
            
            # Calculate status based on hours
            status = record.status
            if hours > 0:
                if hours >= 9:
                    status = 'present'
                elif hours >= 4.5:
                    status = 'half_day'
                else:
                    status = 'absent'
            elif not check_in_time and not check_out_time:
                status = 'absent'
            
            result.append({
                "id": record.id,
                "employee_id": record.employee_id,
                "employee_name": record.employee_name,
                "date": record.date.isoformat(),
                "check_in": check_in_time.isoformat() if check_in_time else None,
                "check_out": check_out_time.isoformat() if check_out_time else None,
                "status": status,
                "hours": round(hours, 2) if hours > 0 else None,
                "remarks": record.remarks
            })
        
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching previous records: {str(e)}")

@router.post("/attendance/modify")
def modify_attendance(
    modify_data: ModifyAttendanceRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Create or modify attendance record"""
    if current_user.role not in ["Admin", "Manager", "HR"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    try:
        target_date = datetime.fromisoformat(modify_data.date).date()
    except:
        raise HTTPException(status_code=400, detail="Invalid date format")
    
    # Get employee
    employee = db.query(User).filter(User.empid == modify_data.employee_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Check if attendance record exists
    attendance = db.query(Attendance).filter(
        and_(
            Attendance.employee_id == modify_data.employee_id,
            Attendance.date == target_date
        )
    ).first()
    
    check_in_time = None
    check_out_time = None
    
    if modify_data.check_in:
        try:
            # Frontend sends full ISO datetime string, parse it directly
            check_in_time = datetime.fromisoformat(modify_data.check_in.replace('Z', '+00:00'))
        except Exception as e:
            print(f"Error parsing check_in: {e}, value: {modify_data.check_in}")
            pass
    
    if modify_data.check_out:
        try:
            # Frontend sends full ISO datetime string, parse it directly
            check_out_time = datetime.fromisoformat(modify_data.check_out.replace('Z', '+00:00'))
        except Exception as e:
            print(f"Error parsing check_out: {e}, value: {modify_data.check_out}")
            pass
    
    # Calculate status based on hours if not provided
    calculated_status = modify_data.status
    hours = 0
    if check_in_time and check_out_time:
        delta = check_out_time - check_in_time
        hours = delta.total_seconds() / 3600
        # Auto-calculate status: >9h = P, 4.5-8.99h = H/D, <4.5h = Abs
        if hours >= 9:
            calculated_status = 'present'
        elif hours >= 4.5:
            calculated_status = 'half_day'
        else:
            calculated_status = 'absent'
    elif not check_in_time and not check_out_time:
        calculated_status = 'absent'
    
    if attendance:
        # Update existing record
        attendance.check_in = check_in_time
        attendance.check_out = check_out_time
        attendance.status = calculated_status
        attendance.remarks = modify_data.remarks
        attendance.updated_at = get_ist_now()
        attendance.hours = hours
    else:
        # Create new record
        attendance = Attendance(
            employee_id=modify_data.employee_id,
            employee_name=employee.name,
            date=target_date,
            check_in=check_in_time,
            check_out=check_out_time,
            status=calculated_status,
            hours=hours,
            remarks=modify_data.remarks
        )
        db.add(attendance)
    
    db.commit()
    db.refresh(attendance)
    
    return {
        "message": "Attendance modified successfully",
        "id": attendance.id
    }

@router.get("/hr/leave-balance")
def get_leave_balance_list(
    year: Optional[int] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get leave balance list for all employees"""
    if current_user.role not in ["Admin", "HR"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    try:
        # Default to current year if not provided
        if not year:
            year = date.today().year
        
        query = db.query(LeaveBalanceList).filter(LeaveBalanceList.year == year)
        
        # Apply search filter
        if search and search.strip():
            search_term = search.strip().lower()
            query = query.filter(
                or_(
                    func.lower(LeaveBalanceList.name).contains(search_term),
                    cast(LeaveBalanceList.empid, String).contains(search_term)
                )
            )
        
        # Limit to 500 records for performance
        records = query.order_by(LeaveBalanceList.name).limit(500).all()
        
        return [
            {
                "id": rec.id,
                "empid": rec.empid,
                "name": rec.name,
                "total_casual_leaves": float(rec.total_casual_leaves) if rec.total_casual_leaves else 0,
                "used_casual_leaves": float(rec.used_casual_leaves) if rec.used_casual_leaves else 0,
                "balance_casual_leaves": float(rec.balance_casual_leaves) if rec.balance_casual_leaves else 0,
                "total_sick_leaves": float(rec.total_sick_leaves) if rec.total_sick_leaves else 0,
                "used_sick_leaves": float(rec.used_sick_leaves) if rec.used_sick_leaves else 0,
                "balance_sick_leaves": float(rec.balance_sick_leaves) if rec.balance_sick_leaves else 0,
                "total_comp_off_leaves": float(rec.total_comp_off_leaves) if rec.total_comp_off_leaves else 0,
                "used_comp_off_leaves": float(rec.used_comp_off_leaves) if rec.used_comp_off_leaves else 0,
                "balance_comp_off_leaves": float(rec.balance_comp_off_leaves) if rec.balance_comp_off_leaves else 0,
                "year": rec.year,
                "updated_by": rec.updated_by,
                "updated_date": rec.updated_date.isoformat() if rec.updated_date else None,
            }
            for rec in records
        ]
    except Exception as e:
        print(f"Error fetching leave balance list: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to fetch leave balance list: {str(e)}")

@router.post("/hr/leave-balance/generate")
def generate_leave_balance(
    year: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Generate leave balance for all employees for a given year"""
    if current_user.role not in ["Admin", "HR"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    try:
        # Default to current year if not provided
        if not year:
            year = date.today().year
        
        # Get current month
        current_month = date.today().month
        
        # Get all employees from users table
        employees = db.query(User).all()
        
        if not employees:
            raise HTTPException(status_code=404, detail="No employees found")
        
        generated_count = 0
        errors = []
        
        for employee in employees:
            try:
                # Convert empid to integer if it's numeric, otherwise use employee.id
                try:
                    empid_value = int(employee.empid) if employee.empid and str(employee.empid).isdigit() else employee.id
                except:
                    empid_value = employee.id
                
                # Calculate leave balances based on current month
                # total_casual_leaves = 12, total_sick_leaves = 12
                total_casual_leaves = 12.0
                total_sick_leaves = 12.0
                
                # balance_casual_leaves = (12 - current_month + 1)
                balance_casual_leaves = float(12 - current_month + 1)
                balance_sick_leaves = float(12 - current_month + 1)
                
                # used_casual_leaves = total_casual_leaves - balance_casual_leaves
                used_casual_leaves = total_casual_leaves - balance_casual_leaves
                used_sick_leaves = total_sick_leaves - balance_sick_leaves
                
                # total_comp_off_leaves = 0
                total_comp_off_leaves = 0.0
                used_comp_off_leaves = 0.0
                balance_comp_off_leaves = 0.0
                
                # Check if record already exists for this employee and year
                existing = db.query(LeaveBalanceList).filter(
                    and_(
                        LeaveBalanceList.empid == empid_value,
                        LeaveBalanceList.year == year
                    )
                ).first()
                
                # Skip if record already exists
                if existing:
                    print(f"  ⊘ Skipped employee {empid_value} ({employee.name}) - Year: {year} (already exists)")
                    continue
                
                # Create new record only if it doesn't exist
                leave_balance_data = {
                    'empid': empid_value,
                    'name': employee.name,
                    'total_casual_leaves': total_casual_leaves,
                    'used_casual_leaves': used_casual_leaves,
                    'balance_casual_leaves': balance_casual_leaves,
                    'total_sick_leaves': total_sick_leaves,
                    'used_sick_leaves': used_sick_leaves,
                    'balance_sick_leaves': balance_sick_leaves,
                    'total_comp_off_leaves': total_comp_off_leaves,
                    'used_comp_off_leaves': used_comp_off_leaves,
                    'balance_comp_off_leaves': balance_comp_off_leaves,
                    'year': year,
                    'updated_by': current_user.name or current_user.empid,
                    'updated_date': get_ist_now()
                }
                
                new_record = LeaveBalanceList(**leave_balance_data)
                db.add(new_record)
                print(f"  ✓ Created leave balance for employee {empid_value} ({employee.name}) - Year: {year}")
                generated_count += 1
            except Exception as e:
                error_msg = f"Error processing employee {employee.empid} ({employee.name if employee else 'Unknown'}): {str(e)}"
                print(error_msg)
                import traceback
                traceback.print_exc()
                errors.append(error_msg)
                continue
        
        if errors:
            print(f"Errors encountered: {errors}")
        
        if generated_count == 0:
            db.rollback()
            error_details = f"No leave balance records were generated. "
            if errors:
                error_details += f"Errors encountered: {', '.join(errors[:3])}"
            raise HTTPException(status_code=400, detail=error_details)
        
        db.commit()
        print(f"Successfully committed {generated_count} leave balance record(s)")
        
        return {
            "message": f"Leave balance generated successfully for {generated_count} employee(s)",
            "generated_count": generated_count,
            "year": year
        }
    
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"Error generating leave balance: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to generate leave balance: {str(e)}")

class AttendanceExcelUploadRequest(BaseModel):
    records: List[dict]

@router.post("/attendance/upload-excel")
def upload_attendance_excel(
    upload_data: AttendanceExcelUploadRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Upload attendance data from Excel and insert into punch_logs table"""
    try:
        if current_user.role not in ['HR', 'Admin']:
            raise HTTPException(status_code=403, detail="Only HR and Admin can upload attendance data")
        
        inserted_count = 0
        errors = []
        
        for record in upload_data.records:
            try:
                employee_id = record.get('employee_id')
                employee_name = record.get('employee_name')
                date_str = record.get('date')
                punchtime_str = record.get('punchtime')
                remarks = record.get('remarks')  # Get remarks from record
                
                if not all([employee_id, date_str, punchtime_str]):
                    errors.append(f"Missing required fields: {record}")
                    continue
                
                # Parse date and time - use naive datetime to avoid timezone issues
                try:
                    # Handle different datetime formats
                    if 'T' in punchtime_str:
                        # ISO format with T - parse and convert to naive
                        dt = datetime.fromisoformat(punchtime_str.replace('Z', '+00:00'))
                        # Convert to naive datetime (remove timezone info)
                        punch_datetime = dt.replace(tzinfo=None)
                    else:
                        # Format: "YYYY-MM-DD HH:MM:SS" or "YYYY-MM-DD HH:MM"
                        # Parse as naive datetime (no timezone)
                        if len(punchtime_str.split(':')) > 2:
                            punch_datetime = datetime.strptime(punchtime_str, '%Y-%m-%d %H:%M:%S')
                        else:
                            punch_datetime = datetime.strptime(punchtime_str, '%Y-%m-%d %H:%M')
                    # Extract date directly from the datetime (no timezone conversion)
                    punch_date = punch_datetime.date()
                except Exception as parse_error:
                    errors.append(f"Invalid date/time format: {punchtime_str} - {str(parse_error)}")
                    continue
                
                # Determine punch type (in or out) based on time
                punch_hour = punch_datetime.hour
                # Assume before 14:00 is IN, after is OUT
                punch_type = 'punch'  # Use default value since we ignore punch_type in calculations
                
                # Create punch log entry
                punch_log = PunchLog(
                    employee_id=str(employee_id),
                    employee_name=employee_name if employee_name else None,
                    date=punch_date,
                    punch_time=punch_datetime,
                    punch_type=punch_type,
                    remarks=remarks if remarks else None
                )
                
                db.add(punch_log)
                inserted_count += 1
            except Exception as e:
                errors.append(f"Error processing record {record}: {str(e)}")
                import traceback
                traceback.print_exc()
                continue
        
        db.commit()
        
        return {
            "success": True,
            "inserted": inserted_count,
            "errors": errors,
            "message": f"Successfully inserted {inserted_count} records"
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"Error uploading attendance Excel: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to upload attendance data: {str(e)}")

@router.post("/attendance/upload-excel-list")
def upload_attendance_list_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Upload attendance list data from Excel file to attendance_list table
    Checks if record exists based on empid, month, and year - updates if exists, inserts if not
    """
    if current_user.role not in ['HR', 'Admin']:
        raise HTTPException(status_code=403, detail="Only HR and Admin can upload attendance data")
    
    try:
        from openpyxl import load_workbook
        from decimal import Decimal
        
        # Read file content
        contents = file.file.read()
        wb = load_workbook(io.BytesIO(contents))
        ws = wb.active
        
        updated_count = 0
        inserted_count = 0
        errors = []
        
        # Skip header rows (rows 1-3 based on export format), start from row 4
        for row_num, row in enumerate(ws.iter_rows(min_row=4, values_only=True), 4):
            try:
                # Skip empty rows
                if not row[0] and not row[1]:
                    continue
                
                # Parse data from Excel columns (matching export format)
                # Columns: EMPLOYEE-NAME, EMP-ID, TOTAL, WORK, W.O, HOLIDAYS, PRESENT, ABSENT, 
                # HALFDAYS, LATE, LOPs, CL, SL, COMP, PAYBLE, MONTH, YEAR, MONTH_NUMBER
                name = str(row[0]).strip() if row[0] else None
                empid_str = str(row[1]).strip() if row[1] else None
                total_days = float(row[2]) if row[2] is not None and str(row[2]).strip() else 0
                working_days = float(row[3]) if row[3] is not None and str(row[3]).strip() else 0
                week_offs = int(row[4]) if row[4] is not None and str(row[4]).strip() else 0
                holi_days = float(row[5]) if row[5] is not None and str(row[5]).strip() else 0
                presents = float(row[6]) if row[6] is not None and str(row[6]).strip() else 0
                absents = float(row[7]) if row[7] is not None and str(row[7]).strip() else 0
                half_days = float(row[8]) if row[8] is not None and str(row[8]).strip() else 0
                late_logs = int(row[9]) if row[9] is not None and str(row[9]).strip() else 0
                lops = float(row[10]) if row[10] is not None and str(row[10]).strip() else 0
                cl = float(row[11]) if row[11] is not None and str(row[11]).strip() else 0
                sl = float(row[12]) if row[12] is not None and str(row[12]).strip() else 0
                comp_offs = float(row[13]) if row[13] is not None and str(row[13]).strip() else 0
                payble_days = float(row[14]) if row[14] is not None and str(row[14]).strip() else 0
                
                # Try to get month/year from new format first (separate columns)
                month = None
                year = None
                if len(row) >= 17:
                    # New format: MONTH (number), YEAR columns
                    month_num_str = str(row[15]).strip() if row[15] else None
                    year_str = str(row[16]).strip() if row[16] else None
                    
                    if month_num_str and year_str:
                        try:
                            month = int(month_num_str)
                            year = int(year_str)
                        except:
                            pass
                
                # Fallback to old format if new format didn't work
                if not month or not year:
                    month_year_str = str(row[15]).strip() if row[15] and len(row) > 15 else None
                
                # Validate required fields
                if not empid_str:
                    errors.append(f"Row {row_num}: Employee ID is required")
                    continue
                
                # Parse empid (convert to integer)
                try:
                    empid_value = int(empid_str)
                except ValueError:
                    errors.append(f"Row {row_num}: Invalid Employee ID format: {empid_str}")
                    continue
                
                    # Parse month and year from month_year_str (format: "December 2025" or "12 2025")
                    if month_year_str:
                        try:
                            # Try parsing "December 2025" format
                            month_year_parts = month_year_str.split()
                            if len(month_year_parts) >= 2:
                                month_name = month_year_parts[0]
                                year_str = month_year_parts[-1]
                                year = int(year_str)
                                # Convert month name to number
                                month_names = ['january', 'february', 'march', 'april', 'may', 'june',
                                             'july', 'august', 'september', 'october', 'november', 'december']
                                month = month_names.index(month_name.lower()) + 1
                        except:
                            try:
                                # Try parsing "12 2025" format
                                parts = month_year_str.split()
                                if len(parts) >= 2:
                                    month = int(parts[0])
                                    year = int(parts[1])
                            except:
                                pass
                
                if not month or not year:
                    errors.append(f"Row {row_num}: Could not parse month and year from: {month_year_str}")
                    continue
                
                # Check if record exists based on empid, month, and year
                existing = db.query(AttendanceList).filter(
                    and_(
                        AttendanceList.empid == empid_value,
                        AttendanceList.month == str(month),
                        AttendanceList.year == year
                    )
                ).first()
                
                # Prepare attendance data (null/empty values become 0)
                attendance_data = {
                    'name': name.upper() if name else None,
                    'empid': empid_value,
                    'total_days': float(total_days) if total_days is not None else 0,
                    'working_days': float(working_days) if working_days is not None else 0,
                    'week_offs': week_offs if week_offs is not None else 0,
                    'holi_days': float(holi_days) if holi_days is not None else 0,
                    'presents': float(presents) if presents is not None else 0,
                    'absents': float(absents) if absents is not None else 0,
                    'half_days': float(half_days) if half_days is not None else 0,
                    'late_logs': late_logs if late_logs is not None else 0,
                    'cl': float(cl) if cl is not None else 0,
                    'sl': float(sl) if sl is not None else 0,
                    'comp_offs': float(comp_offs) if comp_offs is not None else 0,
                    'payble_days': float(payble_days) if payble_days is not None else 0,
                    'lops': float(lops) if lops is not None else 0,
                    'year': year,
                    'month': str(month),
                    'status': 1,
                    'updated_by': current_user.name or current_user.empid,
                    'updated_date': get_ist_now()
                }
                
                if existing:
                    # Update existing record
                    for key, value in attendance_data.items():
                        setattr(existing, key, value)
                    updated_count += 1
                else:
                    # Insert new record
                    new_record = AttendanceList(**attendance_data)
                    db.add(new_record)
                    inserted_count += 1
                    
            except Exception as e:
                errors.append(f"Row {row_num}: Error processing record - {str(e)}")
                import traceback
                traceback.print_exc()
                continue
        
        db.commit()
        
        return {
            "success": True,
            "updated": updated_count,
            "inserted": inserted_count,
            "errors": errors,
            "message": f"Successfully updated {updated_count} record(s) and inserted {inserted_count} record(s)"
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"Error uploading attendance list Excel: {e}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to upload attendance list data: {str(e)}")

@router.get("/attendance/punch-logs")
def get_punch_logs(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    employee_id: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get punch logs for attendance history"""
    try:
        query = db.query(PunchLog)
        
        if employee_id:
            query = query.filter(PunchLog.employee_id == employee_id)
        elif current_user.role == 'Employee':
            query = query.filter(PunchLog.employee_id == current_user.empid)
        
        if start_date:
            try:
                start = datetime.fromisoformat(start_date).date()
                query = query.filter(PunchLog.date >= start)
            except:
                pass
        
        if end_date:
            try:
                end = datetime.fromisoformat(end_date).date()
                query = query.filter(PunchLog.date <= end)
            except:
                pass
        
        punch_logs = query.order_by(PunchLog.date, PunchLog.punch_time).all()
        
        return [
            {
                "id": log.id,
                "employee_id": log.employee_id,
                "employee_name": log.employee_name,
                "date": log.date.isoformat() if log.date else None,
                "punch_time": log.punch_time.isoformat() if log.punch_time else None,
                "punch_type": log.punch_type,
                "image": log.image,
                "location": log.location,
                "remarks": log.remarks
            }
            for log in punch_logs
        ]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to fetch punch logs: {str(e)}")

@router.get("/attendance/punch-logs-by-date")
def get_punch_logs_by_date(
    employee_id: str,
    date: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get punch logs for a specific employee and date"""
    try:
        # Parse date
        try:
            target_date = datetime.fromisoformat(date).date()
        except:
            raise HTTPException(status_code=400, detail="Invalid date format")
        
        # Check permissions - employees can only see their own data
        if current_user.role == 'Employee' and current_user.empid != employee_id:
            raise HTTPException(status_code=403, detail="Access denied")
        
        # Query punch logs
        punch_logs = db.query(PunchLog).filter(
            and_(
                PunchLog.employee_id == employee_id,
                PunchLog.date == target_date
            )
        ).order_by(PunchLog.punch_time).all()
        
        return [
            {
                "id": log.id,
                "employee_id": log.employee_id,
                "employee_name": log.employee_name,
                "date": log.date.isoformat() if log.date else None,
                "punch_time": log.punch_time.isoformat() if log.punch_time else None,
                "punch_type": log.punch_type,
                "image": log.image,
                "location": log.location,
                "remarks": log.remarks
            }
            for log in punch_logs
        ]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to fetch punch logs: {str(e)}")

@router.get("/hr/leaves")
def get_hr_leaves(
    filter: Optional[str] = "all",
    employee_id: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get leave requests for Manager, HR, and Admin roles
    
    - Manager: Shows leaves of employees reporting to them
    - HR: Shows all leaves except logged-in HR person's leaves
    - Admin: Shows all leaves (can see all employees' leaves)
    """
    if current_user.role not in ["Manager", "HR", "Admin"]:
        raise HTTPException(status_code=403, detail="Access denied. Manager, HR, or Admin role required.")
    
    query = db.query(Leave)
    
    if current_user.role == "Manager":
        # Managers can only see leaves of employees reporting to them
        query = query.filter(Leave.report_to == current_user.empid)
    elif current_user.role == "HR":
        # HR can see all leaves except their own
        query = query.filter(Leave.empid != current_user.empid)
    # Admin can see all leaves (no filter needed)
    
    if employee_id:
        query = query.filter(Leave.empid == employee_id)
    
    if filter != "all":
        query = query.filter(Leave.status == filter)
    
    leaves = query.order_by(Leave.applied_date.desc()).all()
    
    return [
        {
            "id": leave.id,
            "employee_id": leave.empid,
            "employee_name": leave.name,
            "applied_date": leave.applied_date.isoformat(),
            "start_date": leave.from_date.isoformat(),
            "end_date": leave.to_date.isoformat(),
            "duration": leave.duration,
            "leave_type": leave.leave_type,
            "report_to": leave.report_to,
            "reason": leave.reason,
            "status": leave.status,
            "approved_by": leave.approved_by,
            "approved_date": leave.approved_date.isoformat() if leave.approved_date else None
        }
        for leave in leaves
    ]

@router.get("/hr/self/leaves")
def get_hr_self_leaves(
    filter: Optional[str] = "all",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get logged-in HR person's own leave requests"""
    if current_user.role != "HR":
        raise HTTPException(status_code=403, detail="Access denied. HR role required.")
    
    query = db.query(Leave).filter(Leave.empid == current_user.empid)
    
    if filter != "all":
        query = query.filter(Leave.status == filter)
    
    leaves = query.order_by(Leave.applied_date.desc()).all()
    
    return [
        {
            "id": leave.id,
            "leave_type": leave.leave_type,
            "start_date": leave.from_date.isoformat(),
            "end_date": leave.to_date.isoformat(),
            "duration": leave.duration,
            "reason": leave.reason,
            "status": leave.status,
            "applied_date": leave.applied_date.isoformat(),
            "approved_by": leave.approved_by,
            "approved_date": leave.approved_date.isoformat() if leave.approved_date else None
        }
        for leave in leaves
    ]

@router.post("/hr/self/leaves")
def create_hr_self_leave(
    leave_data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Create a leave request for logged-in HR person"""
    if current_user.role != "HR":
        raise HTTPException(status_code=403, detail="Access denied. HR role required.")
    
    try:
        from_date = datetime.fromisoformat(leave_data.get("start_date") or leave_data.get("from_date")).date()
        to_date = datetime.fromisoformat(leave_data.get("end_date") or leave_data.get("to_date")).date()
    except:
        raise HTTPException(status_code=400, detail="Invalid date format")
    
    # Calculate duration (all dates included, week-offs and holidays accepted)
    duration = (to_date - from_date).days + 1
    
    new_leave = Leave(
        empid=current_user.empid,
        name=current_user.name,
        applied_date=get_ist_now(),
        from_date=from_date,
        to_date=to_date,
        duration=duration,
        leave_type=leave_data.get("leave_type"),
        report_to=current_user.report_to_id,
        reason=leave_data.get("reason", ""),
        status='pending'
    )
    
    db.add(new_leave)
    db.commit()
    db.refresh(new_leave)
    
    return {
        "message": "Leave request submitted successfully",
        "id": new_leave.id
    }

@router.post("/hr/leaves")
def create_hr_leave(
    leave_data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Create a leave request - HR and Admin roles"""
    if current_user.role not in ["HR", "Admin"]:
        raise HTTPException(status_code=403, detail="Access denied. HR or Admin role required.")
    
    try:
        from_date = datetime.fromisoformat(leave_data.get("start_date") or leave_data.get("from_date")).date()
        to_date = datetime.fromisoformat(leave_data.get("end_date") or leave_data.get("to_date")).date()
    except:
        raise HTTPException(status_code=400, detail="Invalid date format")
    
    employee_id = leave_data.get("employee_id")
    if not employee_id:
        raise HTTPException(status_code=400, detail="employee_id is required")
    
    # Get employee
    employee = db.query(User).filter(User.empid == employee_id).first()
    if not employee:
        raise HTTPException(status_code=404, detail="Employee not found")
    
    # Calculate duration (all dates included, week-offs and holidays accepted)
    duration = (to_date - from_date).days + 1
    
    new_leave = Leave(
        empid=employee_id,
        name=employee.name,
        applied_date=get_ist_now(),
        from_date=from_date,
        to_date=to_date,
        duration=duration,
        leave_type=leave_data.get("leave_type"),
        report_to=employee.report_to_id,
        reason=leave_data.get("reason", ""),
        status=leave_data.get("status", "pending")
    )
    
    db.add(new_leave)
    db.commit()
    db.refresh(new_leave)
    
    return {
        "message": "Leave request created successfully",
        "id": new_leave.id
    }

@router.put("/hr/leaves/{leave_id}")
def update_hr_leave_status(
    leave_id: int,
    status_data: dict = Body(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Update leave status - Manager, HR, and Admin roles"""
    if current_user.role not in ["Manager", "HR", "Admin"]:
        raise HTTPException(status_code=403, detail="Access denied. Manager, HR, or Admin role required.")
    
    leave = db.query(Leave).filter(Leave.id == leave_id).first()
    if not leave:
        raise HTTPException(status_code=404, detail="Leave request not found")
    
    # Check if Manager can approve this leave (only if employee reports to them)
    if current_user.role == "Manager":
        if leave.report_to != current_user.empid:
            raise HTTPException(status_code=403, detail="You can only approve leaves of employees reporting to you.")
    
    # HR and Admin can approve any leave except their own
    if current_user.role in ["HR", "Admin"]:
        if leave.empid == current_user.empid:
            raise HTTPException(status_code=403, detail="You cannot approve your own leave request.")
    
    status = status_data.get("status")
    if status not in ["approved", "rejected"]:
        raise HTTPException(status_code=400, detail="Invalid status")
    
    leave.status = status
    leave.approved_by = current_user.empid
    leave.approved_date = get_ist_now()
    
    db.commit()
    db.refresh(leave)
    
    return {
        "message": f"Leave request {status} successfully",
        "id": leave.id
    }

